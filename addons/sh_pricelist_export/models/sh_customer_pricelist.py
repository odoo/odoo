# JUAN PABLO YAÑEZ CHAPITAL

from odoo import models, fields
import base64
import xlwt
from io import BytesIO
from openpyxl import Workbook

class CustomerPricelistExcelExtended(models.Model):
    _name = "excel.extended"
    _description = 'Customer Pricelist Excel Download'

    excel_file = fields.Binary('Download report Excel')
    file_name = fields.Char('Excel File', size=64)

    def download_report(self):

        return{
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=excel.extended&field=excel_file&download=true&id=%s&filename=%s' % (self.id, self.file_name),
            'target': 'new',
        }


class ShCustomerPricelistWizard(models.Model):
    _name = 'sh.customer.pricelist.wizard'
    _description = 'Sh Customer Pricelist Wizard'

    import_type = fields.Selection([
        ('excel', 'Excel File'),
        ('pdf', 'Pdf File')
    ], default="excel", string="File Type", required=True)

    def print_report(self):
        datas = self.read()[0]
        return self.env.ref('sh_pricelist_export.sh_customer_pricelist_report_action').report_action([], data=datas)
    def action_export_customer_pricelist(self):
        if self:
            #for CSV
            if self.import_type == 'excel':
                workbook = Workbook()                
                workbook.remove(workbook.active)
                # workbook = xlwt.Workbook()
                # bold = xlwt.easyxf(
                #     'font:bold True;pattern: pattern solid, fore_colour gray25;align: horiz left')
                # bold = xlwt.easyxf('font:bold True;align: horiz left')
                # horiz = xlwt.easyxf('align: horiz left')
                # horiz_right = xlwt.easyxf('align: horiz right')
                row = 1
                active_ids = self.env.context.get('active_ids')
                search_partner = self.env['res.partner'].search(
                    [('id', 'in', active_ids)])

                partner_dict = {}

                if search_partner:
                    for partner in search_partner:                        
                        if partner.name not in partner_dict:
                            partner_dict[partner.name] = 1
                            final_partner_name = partner.name
                        else:
                            for key, val in partner_dict.items():
                                if partner.name == key:
                                    final_partner_name = partner.name + ' ' + str(val)
                                    partner_dict[partner.name] = val + 1
                                    break
                        partner_pricelist = partner.property_product_pricelist
                        worksheet = workbook.create_sheet(final_partner_name)
                        # worksheet = workbook.add_sheet(final_partner_name)
                        # worksheet.col(0).width = int(10*260)
                        # worksheet.col(1).width = int(25*260)
                        # worksheet.col(2).width = int(20*260)
                        # worksheet.col(3).width = int(20*260)
                        # worksheet.col(4).width = int(15*260)
                        # worksheet.col(5).width = int(14*260)
                        # worksheet.col(6).width = int(25*260)
                        # worksheet.col(7).width = int(25*260)

                        worksheet.cell(row=1, column=1, value="Partner")
                        worksheet.cell(row=1, column=2, value=partner.name)
                        worksheet.cell(row=1, column=4, value="Pricelist")
                        worksheet.cell(row=1, column=5, value=partner_pricelist.name)

                        # worksheet.write(1, 0, 'Partner', bold)
                        # worksheet.write(1, 1, partner.name)
                        # worksheet.write(1, 4, 'Pricelist', bold)
                        # worksheet.write(1, 5, partner_pricelist.name)

                        worksheet.cell(row=3, column=1, value="ID")
                        worksheet.cell(row=3, column=2, value="Name")
                        worksheet.cell(row=3, column=3, value="Internal Reference")
                        worksheet.cell(row=3, column=4, value="Brand Name")
                        worksheet.cell(row=3, column=5, value="Sale Price")
                        worksheet.cell(row=3, column=6, value="Pricelist Price")
                        worksheet.cell(row=3, column=7, value="Discount(%)")
                        worksheet.cell(row=3, column=8, value="Discount Amount")

                        # worksheet.write(3, 0, "ID", bold)
                        # worksheet.write(3, 1, "Name", bold)
                        # worksheet.write(3, 2, "Internal Reference", bold)
                        # worksheet.write(3, 3, "Brand Name", bold)
                        # worksheet.write(3, 4, "Sale Price", bold)
                        # worksheet.write(3, 5, "Pricelist Price", bold)
                        # worksheet.write(3, 6, "Discount(%)", bold)
                        # worksheet.write(3, 7, "Discount Amount", bold)
                        row = 4
                        product_search = self.env['product.template'].search([
                        ])
                        if product_search:
                            for product in product_search:
                                price_unit = partner_pricelist._compute_price_rule(
                                    [(product, 1.0, partner)], date=fields.Date.today(), uom_id=product.uom_id.id)[product.id][0]
                                discount_amount = 0.00
                                discount = 0.00
                                if product.list_price > price_unit:
                                    discount_amount = product.list_price - price_unit
                                    discount = (
                                        100 * (discount_amount))/product.list_price

                                worksheet.cell(row=row, column=1, value="{0:.0f}".format(product.id) or '')
                                worksheet.cell(row=row, column=2, value=product.name or '')
                                worksheet.cell(row=row, column=3, value=product.default_code or '')
                                worksheet.cell(row=row, column=4, value=product.sh_brand_id.name or '')
                                worksheet.cell(row=row, column=5, value="{0:.2f}".format(product.list_price) or False)
                                worksheet.cell(row=row, column=6, value="{0:.2f}".format(price_unit) or False)
                                worksheet.cell(row=row, column=7, value="{0:.2f}".format(discount) or False)
                                worksheet.cell(row=row, column=8, value="{0:.2f}".format(
                                discount_amount) or False)    
                                # worksheet.write(
                                #     row, 0, product.id or '', horiz)
                                # worksheet.write(row, 1, product.name or '')
                                # worksheet.write(
                                #     row, 2, product.default_code or '')
                                # worksheet.write(
                                #     row, 3, product.sh_brand_id.name or '')
                                # worksheet.write(row, 4, "{0:.2f}".format(
                                #     product.list_price) or False)
                                # worksheet.write(
                                #     row, 5, "{0:.2f}".format(price_unit) or False)
                                # worksheet.write(
                                #     row, 6, "{0:.2f}".format(discount) or False)
                                # worksheet.write(row, 7, "{0:.2f}".format(
                                #     discount_amount) or False)
                                row += 1
                def as_text(value):
                    if value is None:
                        return ""
                    return str(value)
                for column_cells in worksheet.columns:
                    length = max(len(as_text(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length
                filename = ('Customer Pricelist Report' + '.xlsx')
                fp = BytesIO()
                workbook.save(fp)

                export_id = self.env['excel.extended'].sudo().create({
                    'excel_file': base64.encodebytes(fp.getvalue()),
                    'file_name': filename,
                })

                return{
                    'type': 'ir.actions.act_window',
                    'res_id': export_id.id,
                    'res_model': 'excel.extended',
                    'view_type': 'form',
                    'view_mode': 'form',
                    'target': 'new',
                }
