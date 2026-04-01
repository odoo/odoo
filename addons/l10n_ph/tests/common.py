# Part of Odoo. See LICENSE file for full copyright and licensing details.
import io
import unittest

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from odoo.addons.account.tests.common import AccountTestInvoicingCommon


class TestPhCommon(AccountTestInvoicingCommon):
    @classmethod
    @AccountTestInvoicingCommon.setup_country('ph')
    def setUpClass(cls):
        super().setUpClass()

        cls.company_data['company'].write({
            'name': 'Test Company',
            'street': '8 Super Street',
            'city': 'Super City',
            'zip': '8888',
            'country_id': cls.env.ref('base.ph').id,
            'vat': '123-456-789-123',
        })
        cls.partner_a.write({
            'name': 'Test Partner',
            'first_name': 'John',
            'middle_name': 'Doe',
            'last_name': 'Smith',
            'street': '9 Super Street',
            'city': 'Super City',
            'zip': '8888',
            'country_id': cls.env.ref('base.ph').id,
            'vat': '789-456-123-789',
        })
        cls.partner_b.write({
            'name': 'Test Partner Company',
            'property_payment_term_id': cls.pay_terms_a.id,
            'property_supplier_payment_term_id': cls.pay_terms_a.id,
            'street': '10 Super Street',
            'city': 'Super City',
            'zip': '8888',
            'country_id': cls.env.ref('base.ph').id,
            'is_company': True,
            'vat': '789-456-123-456',
        })
        cls.partner_c = cls.env['res.partner'].create({
            'name': 'Test Partner Company Member',
            'property_payment_term_id': cls.pay_terms_a.id,
            'property_supplier_payment_term_id': cls.pay_terms_a.id,
            'property_account_position_id': cls.fiscal_pos_a.id,
            'property_account_receivable_id': cls.company_data['default_account_receivable'].copy().id,
            'property_account_payable_id': cls.company_data['default_account_payable'].copy().id,
            'street': '10 Super Street',
            'city': 'Super City',
            'zip': '8888',
            'country_id': cls.env.ref('base.ph').id,
            'vat': '789-456-123-123',
            'company_id': False,
            'parent_id': cls.partner_b.id,
        })

    def _test_xlsx_file(self, file_content, expected_values):
        """ Takes in the binary content of a xlsx file and a dict of expected values.
        It will then parse the file in order to compare the values with the expected ones.
        The expected values dict format is:
        'row_number': ['cell_1_val', 'cell_2_val', ...]

        :param file_content: The binary content of the xlsx file
        :param expected_values: The dict of expected values
        """
        if load_workbook is None:
            raise unittest.SkipTest("openpyxl not available")

        report_file = io.BytesIO(file_content)
        xlsx = load_workbook(filename=report_file, data_only=True)
        sheet = xlsx.worksheets[0]
        sheet_values = list(sheet.values)

        for row, values in expected_values.items():
            row_values = [v if v is not None else '' for v in sheet_values[row]]
            for row_value, expected_value in zip(row_values, values):
                self.assertEqual(row_value, expected_value)
