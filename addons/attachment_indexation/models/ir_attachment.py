# Part of Odoo. See LICENSE file for full copyright and licensing details.
import io
import importlib.util
import logging
import re
import warnings
import xml.dom.minidom
import zipfile
from lxml import etree

from odoo import api, models
from odoo.tools.lru import LRU

_logger = logging.getLogger(__name__)

if not (importlib.util.find_spec('pdfminer') and importlib.util.find_spec('pdfminer.high_level')):
    _logger.warning("Attachment indexation of PDF documents is unavailable because the 'pdfminer.six' Python library cannot be found on the system. "
                    "You may install it from https://pypi.org/project/pdfminer.six/ (e.g. `pip3 install pdfminer.six`)")

FTYPES = ['docx', 'pptx', 'xlsx', 'opendoc', 'pdf']


index_content_cache = LRU(1)

def textToString(element):
    buff = u""
    for node in element.childNodes:
        if node.nodeType == xml.dom.Node.TEXT_NODE:
            buff += node.nodeValue
        elif node.nodeType == xml.dom.Node.ELEMENT_NODE:
            buff += textToString(node)
    return buff


def _clean_text_content(buf):
    """Clean PDF content: remove NULs, normalize whitespace and line breaks."""
    if not buf:
        return buf
    # Remove NULs, normalize CRLF/CR to LF, replace tabs with spaces
    buf = buf.translate({
        ord('\x00'): None,
        ord('\r'): None,
        ord('\t'): ord(' '),
    })

    # Collapse runs of whitespace while preserving at most a single blank line
    def _compact_whitespace(match):
        chunk = match.group(0)
        newline_count = chunk.count('\n')
        if newline_count == 0:
            return ' '
        return '\n\n' if newline_count > 1 else '\n'

    buf = re.sub(r'\s{2,}', _compact_whitespace, buf)
    return buf.strip()


def _csv_escape(value):
    if value is None:
        return ''
    value = str(value)
    if ',' in value or '"' in value or '\n' in value or '\r' in value:
        return '"' + value.replace('"', '""') + '"'
    return value


class IrAttachment(models.Model):
    _inherit = 'ir.attachment'

    def _index_docx(self, bin_data):
        '''Index Microsoft .docx documents'''
        buf = u""
        f = io.BytesIO(bin_data)
        if zipfile.is_zipfile(f):
            try:
                zf = zipfile.ZipFile(f)
                content = xml.dom.minidom.parseString(zf.read("word/document.xml"))
                for val in ["w:p", "w:h", "text:list"]:
                    for element in content.getElementsByTagName(val):
                        buf += textToString(element) + "\n"
            except Exception:
                pass
        return buf

    def _index_pptx(self, bin_data):
        '''Index Microsoft .pptx documents'''

        buf = u""
        f = io.BytesIO(bin_data)
        if zipfile.is_zipfile(f):
            try:
                zf = zipfile.ZipFile(f)
                zf_filelist = [x for x in zf.namelist() if x.startswith('ppt/slides/slide')]
                for i in range(1, len(zf_filelist) + 1):
                    content = xml.dom.minidom.parseString(zf.read('ppt/slides/slide%s.xml' % i))
                    for val in ["a:t"]:
                        for element in content.getElementsByTagName(val):
                            buf += textToString(element) + "\n"
            except Exception:
                pass
        return buf

    def _index_xlsx(self, bin_data):
        '''Index Microsoft .xlsx documents'''

        try:
            from openpyxl import load_workbook  # noqa: PLC0415
            logging.getLogger("openpyxl").setLevel(logging.CRITICAL)
        except ImportError:
            _logger.info('openpyxl is not installed.')
            return ""

        f = io.BytesIO(bin_data)
        all_sheets = []
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = load_workbook(f, data_only=True, read_only=True)
                for sheet in workbook.worksheets:
                    sheet_name = sheet.title
                    sheet_name_escaped = _csv_escape(sheet_name)
                    sheet_rows = []
                    for row in sheet.iter_rows(values_only=True):
                        if not any(row):
                            continue
                        row_cells = [sheet_name_escaped] + [
                            _csv_escape(str(cell) if cell is not None else '') for cell in row
                        ]
                        sheet_rows.append(','.join(row_cells))
                    sheet_data = '\n'.join(sheet_rows)
                    if sheet_data:
                        all_sheets.append(sheet_data)
        except Exception:  # noqa: BLE001
            pass

        all_sheets_str = '\n\n'.join(all_sheets)
        return _clean_text_content(all_sheets_str)

    def _index_opendoc(self, bin_data):
        '''Index OpenDocument documents (.odt, .ods...)'''

        f = io.BytesIO(bin_data)
        buf = []
        MAX_COLUMN_REPEAT = 100
        MAX_ROW_REPEAT = 50
        main_namespaces = {
            'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
            'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0',
            'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
            'manifest': 'urn:oasis:names:tc:opendocument:xmlns:manifest:1.0'
        }

        def extract_row(row):
            cells = []
            for cell in row.xpath('.//table:table-cell | .//table:covered-table-cell', namespaces=main_namespaces):
                repeat = cell.get(f'{{{main_namespaces["table"]}}}number-columns-repeated')
                repeat_count = min(int(repeat), MAX_COLUMN_REPEAT) if repeat and repeat.isdigit() else 1
                text_parts = cell.xpath('.//text:p//text()', namespaces=main_namespaces)
                cell_text = ' '.join(t.strip() for t in text_parts if t.strip())
                cells.extend([cell_text] * repeat_count)
            return cells

        def extract_spreadsheet(content):
            sheets_csv = []
            tables = content.xpath('.//table:table', namespaces=main_namespaces)
            for table in tables:
                table_rows = []
                table_name = table.get(f'{{{main_namespaces["table"]}}}name')
                if not table_name:
                    table_name = f"Sheet{len(sheets_csv) + 1}"
                table_name_escaped = _csv_escape(table_name)
                for row in table.xpath('.//table:table-row', namespaces=main_namespaces):
                    row_repeat = row.get(f'{{{main_namespaces["table"]}}}number-rows-repeated')
                    row_repeat_count = min(int(row_repeat), MAX_ROW_REPEAT) if row_repeat and row_repeat.isdigit() else 1

                    cells = extract_row(row)
                    if not any(cells):
                        continue

                    while cells and not cells[-1]:
                        cells.pop()

                    row_str = ','.join([table_name_escaped] + list(map(_csv_escape, cells)))
                    if row_str.replace(',', '').strip():
                        table_rows.extend([row_str] * row_repeat_count)

                if table_rows:
                    sheets_csv.append('\n'.join(table_rows))

            return sheets_csv

        def extract_text(content):
            lines = []
            for element in content.xpath('.//text:p | .//text:h | .//text:list-item', namespaces=main_namespaces):
                text = ''.join(element.xpath('.//text()', namespaces=main_namespaces)).strip()
                if text:
                    lines.append(text)
            return lines

        if zipfile.is_zipfile(f):
            try:
                zf = zipfile.ZipFile(f)
                content = etree.fromstring(zf.read('content.xml'))
                mime_type = zf.read('mimetype').decode('utf-8').strip()
                if mime_type and 'spreadsheet' in mime_type:
                    buf.extend(extract_spreadsheet(content))
                else:
                    buf.extend(extract_text(content))
            except Exception:
                pass

        buf_str = '\n\n'.join(buf)
        return _clean_text_content(buf_str)

    def _index_pdf(self, bin_data):
        '''Index PDF documents'''
        if not bin_data.startswith(b'%PDF-'):
            return ""
        try:
            if not importlib.util.find_spec('pdfminer.high_level'):
                return ""
            from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter  # noqa: PLC0415
            from pdfminer.converter import TextConverter  # noqa: PLC0415
            from pdfminer.layout import LAParams  # noqa: PLC0415
            from pdfminer.pdfpage import PDFPage  # noqa: PLC0415
            logging.getLogger("pdfminer").setLevel(logging.CRITICAL)
        except ImportError:
            # warned already during init of module
            return ""
        f = io.BytesIO(bin_data)
        try:
            resource_manager = PDFResourceManager()
            laparams = LAParams(detect_vertical=True)

            with io.StringIO() as content, TextConverter(
                resource_manager,
                content,
                laparams=laparams
            ) as device:
                interpreter = PDFPageInterpreter(resource_manager, device)
                for page in PDFPage.get_pages(f):
                    interpreter.process_page(page)

                buf = content.getvalue()
            return _clean_text_content(buf)
        except Exception:  # noqa: BLE001
            return ""

    @api.model
    def _index(self, bin_data, mimetype, checksum=None):
        if checksum:
            cached_content = index_content_cache.get(checksum)
            if cached_content:
                return cached_content
        res = False
        for ftype in FTYPES:
            buf = getattr(self, '_index_%s' % ftype)(bin_data)
            if buf:
                res = buf.replace('\x00', '')
                break

        res = res or super(IrAttachment, self)._index(bin_data, mimetype, checksum=checksum)
        if checksum:
            index_content_cache[checksum] = res
        return res

    def copy(self, default=None):
        for attachment in self:
            index_content_cache[attachment.checksum] = attachment.index_content
        return super().copy(default=default)
