import base64
import csv
import io

from odoo import fields, models
from odoo.exceptions import UserError, ValidationError

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


class GovAuditoriaEspelhoImportWizard(models.TransientModel):
    _name = "gov.auditoria.espelho.import.wizard"
    _description = "Mirror Import Wizard"

    ciclo_id = fields.Many2one("gov.auditoria.ciclo", required=True, ondelete="cascade")
    upload_file = fields.Binary(required=True)
    upload_filename = fields.Char(required=True)
    delimiter = fields.Char(default=";")
    has_header = fields.Boolean(default=True)
    preview_text = fields.Text(readonly=True)

    _HEADER_MAP = {
        "tipo_movimento": "tipo_movimento",
        "tipo": "tipo_movimento",
        "data_movimento": "data_movimento",
        "data": "data_movimento",
        "valor": "valor",
        "historico": "historico",
        "descricao": "historico",
        "conta_codigo": "conta_codigo",
        "fonte_recurso": "fonte_recurso",
        "natureza_despesa": "natureza_despesa",
        "funcional": "funcional",
    }

    def _read_rows(self):
        self.ensure_one()
        raw = base64.b64decode(self.upload_file or b"")
        if not raw:
            raise ValidationError("Arquivo de importacao vazio.")
        filename = (self.upload_filename or "").lower()
        if filename.endswith(".csv"):
            text = raw.decode("utf-8-sig")
            return list(csv.DictReader(io.StringIO(text), delimiter=self.delimiter or ";"))
        if filename.endswith(".xlsx"):
            if load_workbook is None:
                raise UserError("openpyxl nao esta disponivel para importar XLSX.")
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            header = [str(col or "").strip() for col in rows[0]]
            data_rows = rows[1:] if self.has_header else rows
            result = []
            for row in data_rows:
                result.append({header[index]: row[index] for index in range(min(len(header), len(row)))})
            return result
        raise UserError("Formato nao suportado. Use CSV ou XLSX.")

    def _normalize_row(self, row):
        normalized = {}
        for key, value in row.items():
            mapped = self._HEADER_MAP.get((key or "").strip().lower())
            if mapped:
                normalized[mapped] = value
        normalized.setdefault("origem", "importado_csv")
        return normalized

    def action_preview(self):
        self.ensure_one()
        rows = self._read_rows()[:5]
        lines = []
        for row in rows:
            lines.append(str(self._normalize_row(row)))
        self.preview_text = "\n".join(lines) if lines else "Sem linhas para visualizar."
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_import(self):
        self.ensure_one()
        rows = self._read_rows()
        if not rows:
            raise UserError("Nao ha linhas para importar.")

        attachment = self.env["ir.attachment"].create(
            {
                "name": self.upload_filename,
                "datas": self.upload_file,
                "res_model": "gov.auditoria.ciclo",
                "res_id": self.ciclo_id.id,
            }
        )

        create_vals = []
        for raw_row in rows:
            row = self._normalize_row(raw_row)
            if not row.get("tipo_movimento") or row.get("valor") in (None, "") or not row.get("historico"):
                raise ValidationError(f"Linha invalida para importacao: {raw_row}")
            create_vals.append(
                {
                    "ciclo_id": self.ciclo_id.id,
                    "tipo_movimento": str(row["tipo_movimento"]).strip().lower(),
                    "data_movimento": row.get("data_movimento") or fields.Date.today(),
                    "valor": float(row["valor"]),
                    "historico": str(row["historico"]),
                    "origem": "importado_csv",
                    "documento_fonte_id": attachment.id,
                    "hash_fonte": attachment.checksum,
                    "conta_codigo": row.get("conta_codigo") or False,
                    "fonte_recurso": row.get("fonte_recurso") or False,
                    "natureza_despesa": row.get("natureza_despesa") or False,
                    "funcional": row.get("funcional") or False,
                }
            )
        self.env["gov.auditoria.espelho"].create(create_vals)
        return {"type": "ir.actions.act_window_close"}
