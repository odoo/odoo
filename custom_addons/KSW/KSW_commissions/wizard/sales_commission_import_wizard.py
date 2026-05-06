"""KSW Sales/Collection Commission Import Wizard.

Reads the accountant's monthly Excel reports and populates
``ksw.sales.commission.line.achieved_sales`` /
``ksw.sales.commission.line.achieved_collection`` /
``ksw.sales.commission.line.target_collection`` automatically.

Expected file formats (column index = 0-based)
----------------------------------------------
**Sales file** (e.g. ``Sales March 2026.xlsx``):
  Col 0: Account number
  Col 1: Customer name
  Col 2: Sales amount (VAT-inclusive)         — ignored
  Col 3: Pre-tax sales (المبيعات قبل الضريبة) → ``achieved_sales``
  Col 4: Salesman name (البائع)               → employee key

**Collection file** (e.g. ``Collection March 2026.xlsx``):
  Col 0: Account number
  Col 1: Customer name
  Col 2: Balance                              — ignored
  Col 3: Aging                                — ignored
  Col 4: Target                               → ``target_collection``
  Col 5: Amount collected (المحصل)            → ``achieved_collection``
  Col 6: Collection rep name (مندوب التحصيل)  → employee key
  Col 7: Collection %                         — ignored

Row detection
-------------
A row is considered a data row when col 0 is non-empty AND the name
column for that file is a non-empty string. Subtotal/grand-total rows
(blank account number) and ``#DIV/0!`` rows are skipped automatically.
A bad numeric cell logs a warning instead of aborting the import.

Name matching
-------------
For each name found in the file the wizard looks for an ``hr.employee``
record where ``x_commission_import_name`` equals the name (exact,
case-insensitive).  If not found, it falls back to matching on
``hr.employee.name``.  Unmatched names are listed in the chatter.
"""
import base64
import io
import logging

from markupsafe import Markup, escape

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


class KswSalesCommissionImportWizard(models.TransientModel):
    _name = 'ksw.sales.commission.import.wizard'
    _description = 'KSW Sales/Collection Commission Import Wizard'

    sheet_id = fields.Many2one(
        'ksw.sales.commission.sheet',
        required=True, ondelete='cascade',
        string='Commission Sheet',
    )
    sales_file = fields.Binary(string='Sales Excel File')
    sales_filename = fields.Char(string='Sales File Name')
    collection_file = fields.Binary(string='Collection Excel File')
    collection_filename = fields.Char(string='Collection File Name')
    collection_exclude_vat = fields.Boolean(
        string='Exclude VAT (÷ 1.15)',
        default=False,
        help='When checked, both the Target and Collected amounts from the '
             'collection file are divided by 1.15 to remove the 15% VAT '
             'before being written to the commission line.',
    )

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('openpyxl is required for Excel import.'))
        if not self.sales_file and not self.collection_file:
            raise UserError(_(
                'Please upload at least one Excel file '
                '(Sales and/or Collection).'))

        sheet = self.sheet_id
        if sheet.state != 'draft':
            raise UserError(_(
                'The commission sheet must be in draft state to import data.'))

        # -- parse files -------------------------------------------------
        # sales_data: {salesman_lower: {'total': float, 'by_customer': {cust_lower: float}}}
        # collection_totals: {rep_lower: {'collected': float, 'target': float}}
        sales_data = {}
        collection_totals = {}

        if self.sales_file:
            sales_data = self._parse_sales_file(self.sales_file)
            if not sales_data:
                raise UserError(_(
                    'No data rows could be read from the Sales file. '
                    'Check that the column layout matches the expected '
                    'format (Account · Customer · Sales · Pre-tax · '
                    'Salesman).'))
        if self.collection_file:
            collection_totals = self._parse_collection_file(
                self.collection_file,
                exclude_vat=self.collection_exclude_vat)
            if not collection_totals:
                raise UserError(_(
                    'No data rows could be read from the Collection file. '
                    'Check that the column layout matches the expected '
                    'format (Account · Customer · Balance · Aging · '
                    'Target · Collected · Rep · %).'))

        # -- build employee map ------------------------------------------
        all_names_lower = set(sales_data) | set(collection_totals)
        emp_map = self._build_employee_map(all_names_lower)

        # -- upsert lines ------------------------------------------------
        Line = self.env['ksw.sales.commission.line']
        Profile = self.env['ksw.salesperson.profile']
        imported_lines = []
        unmatched = sorted(n for n in all_names_lower if n not in emp_map)
        sheet_year = (
            fields.Date.to_date(sheet.period).year if sheet.period else None
        )

        for name_lower in sorted(all_names_lower):
            employee = emp_map.get(name_lower)
            if not employee:
                continue

            emp_sales = sales_data.get(name_lower)   # dict or None
            coll_data = collection_totals.get(name_lower)  # dict or None
            coll_amt = coll_data['collected'] if coll_data else None
            target_amt = coll_data['target'] if coll_data else None

            # ----------------------------------------------------------
            # Fetch profile splits for this employee / year
            # ----------------------------------------------------------
            splits = self.env['ksw.salesperson.profile.client.split']
            if sheet_year:
                profile = Profile.sudo().search([
                    ('employee_id', '=', employee.id),
                    ('year', '=', sheet_year),
                    ('active', '=', True),
                ], limit=1)
                if profile:
                    splits = profile.split_ids

            # ----------------------------------------------------------
            # Handle split lines (each covers a named client bucket)
            # ----------------------------------------------------------
            for split in splits:
                # Build lookup sets + reverse map for chatter detail.
                # Priority per partner:
                #   1. x_client_account_number (most stable — col 0 of Excel)
                #   2. x_commission_import_name (name alias — col 1)
                #   3. partner.name             (last resort — col 1)
                acc_keys = {}    # acc_lower → partner display label
                name_keys = {}   # name_lower → partner display label
                for p in split.rule_id.partner_ids:
                    acc = (p.x_client_account_number or '').strip().lower()
                    alias = (p.x_commission_import_name or '').strip().lower()
                    pname = (p.name or '').strip().lower()
                    # Human-readable label: "AccNo — Partner Name" or just name
                    if p.x_client_account_number:
                        label = f"{p.x_client_account_number} — {p.name or ''}"
                    else:
                        label = p.name or '?'
                    if acc:
                        acc_keys[acc] = label
                    elif alias:
                        name_keys[alias] = label
                    elif pname:
                        name_keys[pname] = label

                # Sum matching customers' sales — track per-client detail
                split_sales = 0.0
                matched_detail = []   # [(label, amount), ...]
                unmatched_partners = [
                    p.x_client_account_number or p.name
                    for p in split.rule_id.partner_ids
                ]
                if emp_sales:
                    for acc_lower, amt in emp_sales['by_account'].items():
                        if acc_lower in acc_keys:
                            split_sales += amt
                            lbl = acc_keys[acc_lower]
                            matched_detail.append((lbl, amt))
                            if lbl in unmatched_partners:
                                unmatched_partners.remove(lbl)
                    for cust_lower, amt in emp_sales['by_customer'].items():
                        if cust_lower in name_keys:
                            split_sales += amt
                            lbl = name_keys[cust_lower]
                            matched_detail.append((lbl, amt))
                            if lbl in unmatched_partners:
                                unmatched_partners.remove(lbl)

                # Find or create the split line
                existing_split = Line.search([
                    ('sheet_id', '=', sheet.id),
                    ('employee_id', '=', employee.id),
                    ('split_id', '=', split.id),
                ], limit=1)
                split_vals = {}
                if emp_sales is not None:
                    split_vals['achieved_sales'] = split_sales

                if existing_split:
                    if split_vals:
                        existing_split.write(split_vals)
                    status = 'updated (split)'
                else:
                    new_line = Line.create({
                        'sheet_id': sheet.id,
                        'employee_id': employee.id,
                        'split_id': split.id,
                    })
                    if split_vals:
                        new_line.write(split_vals)
                    status = 'created (split)'
                imported_lines.append((
                    f"{employee.display_name} [{split.label}]",
                    split_sales if emp_sales is not None else None,
                    None, None, status,
                    matched_detail, unmatched_partners,
                ))

            # ----------------------------------------------------------
            # General line — receives the FULL total sales (not reduced)
            # and all collection data.  The split lines calculate extra
            # commission on their client subset independently.
            # ----------------------------------------------------------
            general_sales = None
            if emp_sales is not None:
                general_sales = emp_sales['total']

            existing_general = Line.search([
                ('sheet_id', '=', sheet.id),
                ('employee_id', '=', employee.id),
                ('split_id', '=', False),
            ], limit=1)

            gen_vals = {}
            if general_sales is not None:
                gen_vals['achieved_sales'] = general_sales
            if coll_amt is not None:
                gen_vals['achieved_collection'] = coll_amt
            if target_amt is not None:
                gen_vals['target_collection'] = target_amt

            if gen_vals:
                if existing_general:
                    existing_general.write(gen_vals)
                    imported_lines.append((
                        employee.display_name, general_sales,
                        coll_amt, target_amt, 'updated', [], []))
                else:
                    new_line = Line.create({
                        'sheet_id': sheet.id,
                        'employee_id': employee.id,
                    })
                    new_line.write(gen_vals)
                    imported_lines.append((
                        employee.display_name, general_sales,
                        coll_amt, target_amt, 'created', [], []))

        # -- Collection Manager: total-collection pass -------------------
        # Employees whose profile has x_collection_based_on_total=True
        # receive the grand total of ALL collections (and targets) from
        # the collection file, regardless of what name appears in the file.
        if self.collection_file and sheet_year:
            grand_collected = sum(
                v['collected'] for v in collection_totals.values()
            )
            grand_target = sum(
                v['target'] for v in collection_totals.values()
            )
            mgr_profiles = Profile.sudo().search([
                ('year', '=', sheet_year),
                ('active', '=', True),
                ('x_collection_based_on_total', '=', True),
            ])
            for profile in mgr_profiles:
                employee = profile.employee_id
                existing = Line.search([
                    ('sheet_id', '=', sheet.id),
                    ('employee_id', '=', employee.id),
                    ('split_id', '=', False),
                ], limit=1)
                mgr_vals = {
                    'achieved_collection': grand_collected,
                    'target_collection': grand_target,
                }
                if existing:
                    existing.write(mgr_vals)
                    action_lbl = 'updated'
                else:
                    new_line = Line.create({
                        'sheet_id': sheet.id,
                        'employee_id': employee.id,
                    })
                    new_line.write(mgr_vals)
                    action_lbl = 'created'
                imported_lines.append((
                    f"{employee.display_name} [Total Collection]",
                    None, grand_collected, grand_target,
                    action_lbl, [], [],
                ))

        # -- chatter summary ----------------------------------------------
        self._post_import_summary(sheet, imported_lines, unmatched)

        return {
            'type': 'ir.actions.act_window',
            'res_id': sheet.id,
            'view_mode': 'form',
            'target': 'current',
        }

    # ------------------------------------------------------------------
    # Parsers
    # ------------------------------------------------------------------
    @staticmethod
    def _is_data_row(row, name_col):
        """Return True for a genuine data row.

        A data row has a non-empty col 0 (account number) AND a
        non-empty string in the name column. Subtotals, blanks, and
        ``#DIV/0!`` rows fail one of these conditions.
        """
        if len(row) <= name_col:
            return False
        acc = row[0]
        if acc is None:
            return False
        if isinstance(acc, str) and not acc.strip():
            return False
        name = row[name_col]
        if name is None:
            return False
        if not isinstance(name, str):
            return False
        return name.strip() != ''

    @staticmethod
    def _safe_float(value):
        """Coerce a cell value to float; return None if not coercible.

        Accepts int/float as-is, strips commas/spaces from strings,
        treats ``#DIV/0!`` and other error markers as None.
        """
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip().replace(',', '').replace(' ', '')
            if not cleaned or cleaned.startswith('#'):
                return None
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

    def _parse_sales_file(self, file_data):
        """Return per-salesman sales data from the Excel file.

        Return format::

            {
                salesman_name_lower: {
                    'total': float,
                    'by_customer': {customer_name_lower: float},
                    'by_account':  {account_number_lower: float},
                },
                ...
            }

        ``by_account`` uses col 0 (account number) as the key —
        this is the most reliable matching key for split buckets.
        ``by_customer`` uses col 1 (customer name) as a fallback.

        See module docstring for the column contract.
        """
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(file_data)), data_only=True)
        ws = wb.active
        result = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if not self._is_data_row(row, name_col=4):
                continue
            salesman = str(row[4]).strip()
            account = str(row[0]).strip() if row[0] is not None else ''
            customer = str(row[1]).strip() if row[1] is not None else ''
            amount = self._safe_float(row[3])
            if amount is None:
                _logger.warning(
                    'KSW import: skipping non-numeric pre-tax sales '
                    'in row %d (salesman=%r, value=%r)',
                    i + 1, salesman, row[3])
                amount = 0.0
            key = salesman.lower()
            bucket = result.setdefault(
                key, {'total': 0.0, 'by_customer': {}, 'by_account': {}})
            bucket['total'] += amount
            if account:
                acc_key = account.lower()
                bucket['by_account'][acc_key] = (
                    bucket['by_account'].get(acc_key, 0.0) + amount
                )
            if customer:
                cust_key = customer.lower()
                bucket['by_customer'][cust_key] = (
                    bucket['by_customer'].get(cust_key, 0.0) + amount
                )
        return result

    def _parse_collection_file(self, file_data, exclude_vat=False):
        """Return ``{rep_name_lower: {'collected': float, 'target': float}}``
        from the collection Excel file.

        When ``exclude_vat=True`` both Target and Collected amounts are
        divided by 1.15 to strip the 15% VAT before being returned.

        See module docstring for the column contract.
        """
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(file_data)), data_only=True)
        ws = wb.active
        vat_divisor = 1.15 if exclude_vat else 1.0
        totals = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if not self._is_data_row(row, name_col=6):
                continue
            rep = str(row[6]).strip()
            collected = self._safe_float(row[5])
            if collected is None:
                _logger.warning(
                    'KSW import: skipping non-numeric collected amount '
                    'in row %d (rep=%r, value=%r)',
                    i + 1, rep, row[5])
                collected = 0.0
            target = self._safe_float(row[4])
            if target is None:
                target = 0.0
            key = rep.lower()
            bucket = totals.setdefault(
                key, {'collected': 0.0, 'target': 0.0})
            bucket['collected'] += collected / vat_divisor
            bucket['target'] += target / vat_divisor
        return totals

    # ------------------------------------------------------------------
    # Employee matching
    # ------------------------------------------------------------------
    def _build_employee_map(self, names_lower):
        """Return {name_lower: hr.employee} for as many names as possible.

        Priority:
          1. ``x_commission_import_name.lower()`` exact match
          2. ``hr.employee.name.lower()`` exact match
        """
        employees = self.env['hr.employee'].sudo().search([
            ('active', '=', True),
        ])
        result = {}

        # Build lookup by import alias first (highest priority).
        alias_map = {}
        for emp in employees:
            alias = (emp.x_commission_import_name or '').strip().lower()
            if alias:
                alias_map[alias] = emp

        # Build lookup by employee name.
        name_map = {}
        for emp in employees:
            name_map[emp.name.strip().lower()] = emp

        for name_lower in names_lower:
            if name_lower in alias_map:
                result[name_lower] = alias_map[name_lower]
            elif name_lower in name_map:
                result[name_lower] = name_map[name_lower]
            # else: unmatched — caller will log a warning

        return result

    # ------------------------------------------------------------------
    # Chatter summary
    # ------------------------------------------------------------------
    def _post_import_summary(self, sheet, imported_lines, unmatched):
        """Post a chatter note with a full import summary."""
        lines_html = Markup('')
        for entry in imported_lines:
            emp_name, sales, coll, target, action = entry[:5]
            matched_detail = entry[5] if len(entry) > 5 else []
            unmatched_partners = entry[6] if len(entry) > 6 else []

            s_str = f'SAR {sales:,.2f}' if sales is not None else '—'
            c_str = f'SAR {coll:,.2f}' if coll is not None else '—'
            t_str = f'SAR {target:,.2f}' if target is not None else '—'

            # Per-client breakdown for split lines (skip zero-amount rows)
            detail_html = Markup('')
            active_rows = [(lbl, amt) for lbl, amt in matched_detail if amt]
            if active_rows:
                rows = Markup('').join(
                    Markup(
                        '<tr>'
                        '<td style="padding:2px 10px;">{lbl}</td>'
                        '<td style="padding:2px 10px;text-align:right;'
                        'font-family:monospace;">SAR {amt}</td>'
                        '</tr>'
                    ).format(lbl=lbl, amt=f'{amt:,.2f}')
                    for lbl, amt in sorted(active_rows, key=lambda x: -x[1])
                )
                detail_html += Markup(
                    '<table style="margin:4px 0 4px 16px;font-size:0.9em;'
                    'border-collapse:collapse;">'
                    '<thead><tr style="border-bottom:1px solid #ccc;">'
                    '<th style="text-align:left;padding:2px 10px;">'
                    'Client</th>'
                    '<th style="text-align:right;padding:2px 10px;">'
                    'Amount</th>'
                    '</tr></thead>'
                    '<tbody>{rows}</tbody>'
                    '</table>'
                ).format(rows=rows)
            if unmatched_partners:
                missing = Markup(', ').join(
                    escape(str(x)) for x in unmatched_partners
                )
                detail_html += Markup(
                    '<div style="margin-left:16px;font-size:0.85em;'
                    'color:#c0392b;">⚠ Not found in Excel: {m}</div>'
                ).format(m=missing)

            lines_html += Markup(
                '<li><b>{name}</b> [{action}] '
                'Sales: {s} | Target Coll: {t} | Collected: {c}'
                '{detail}</li>'
            ).format(
                name=escape(emp_name), action=escape(action),
                s=s_str, t=t_str, c=c_str,
                detail=detail_html,
            )

        warn_html = Markup('')
        if unmatched:
            names = Markup('').join(
                Markup('<li>{n}</li>').format(n=escape(n)) for n in unmatched
            )
            warn_html = Markup(
                '<br/><b>⚠ Unmatched salesman names (no employee found):</b>'
                '<ul>{names}</ul>'
                'Set the <i>Commission Import Name</i> field on the '
                'corresponding employee records to fix the mapping.'
            ).format(names=names)

        files_info = []
        if self.sales_filename:
            files_info.append(
                Markup('<b>Sales:</b> {f}').format(f=escape(self.sales_filename))
            )
        if self.collection_filename:
            vat_note = (
                Markup(' <i>(VAT excluded ÷1.15)</i>')
                if self.collection_exclude_vat else Markup('')
            )
            files_info.append(
                Markup('<b>Collection:</b> {f}{v}').format(
                    f=escape(self.collection_filename), v=vat_note)
            )
        files_str = Markup(' | ').join(files_info) if files_info else Markup('—')

        body = Markup(
            '<b>📥 Excel Import completed</b> ({files})<br/>'
            '<b>{n} line(s) updated/created:</b>'
            '<ul>{lines}</ul>'
            '{warn}'
        ).format(
            files=files_str,
            n=len(imported_lines),
            lines=lines_html,
            warn=warn_html,
        )
        sheet.message_post(body=body, subtype_xmlid='mail.mt_note')






