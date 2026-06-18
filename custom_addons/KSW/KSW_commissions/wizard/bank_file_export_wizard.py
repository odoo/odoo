"""Bank-file export wizard for KSW Commission batches.

Generates WPS Excel / Kawthar TXT files from the ``total_payable`` on
each ``ksw.commission.sheet`` in the batch. Reuses the same file formats
as ``KSW_payroll`` but the data source is the commission sheet totals
(not payslip lines).
"""
import base64
import io
import zipfile

from odoo import _, api, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None


EXPORT_MODES = [
    ('all_excel',    'All banks – Excel files'),
    ('all_txt',      'All banks – Text files (Kawthar)'),
    ('specific_excel', 'Specific bank – Excel'),
    ('specific_txt', 'Specific bank – Text file'),
]


class KswCommissionBankExportWizard(models.TransientModel):
    _name = 'ksw.commission.bank.export.wizard'
    _description = 'KSW Commission Bank File Export Wizard'

    batch_id = fields.Many2one(
        'ksw.commission.batch', required=True, readonly=True,
    )
    export_mode = fields.Selection(
        EXPORT_MODES, required=True, default='all_excel',
    )
    bank_account_id = fields.Many2one(
        'res.partner.bank',
        domain="[('x_file_type', '!=', False),"
               " ('partner_id', '=', company_partner_id)]",
    )
    company_partner_id = fields.Many2one(
        'res.partner', compute='_compute_company_partner',
    )
    value_date = fields.Date(
        default=fields.Date.context_today,
        help='Payment value date used in TXT files.',
    )
    operation_code = fields.Selection(
        [('1', '1 – New'), ('2', '2 – Renewal'), ('3', '3 – Delete')],
        default='2', string='Kawthar Operation',
    )

    @api.depends('batch_id')
    def _compute_company_partner(self):
        for rec in self:
            rec.company_partner_id = rec.env.company.partner_id

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _group_and_validate(self, require_type=None):
        batch = self.batch_id
        if not batch.sheet_ids:
            raise UserError(_('No sheets in this batch to export.'))
        groups = batch._group_sheets_by_bank_account()
        no_bank = groups.pop(self.env['res.partner.bank'], None)
        if no_bank:
            names = ', '.join(no_bank.mapped('employee_id.name'))
            raise UserError(_(
                'The following employees have no bank account and no '
                'batch-level fallback:\n%s', names))
        no_type = [b for b in groups if not b.x_file_type]
        if no_type:
            accs = ', '.join(b.acc_number or str(b.id) for b in no_type)
            raise UserError(_(
                'These bank accounts have no Payroll File Type set:\n%s', accs))
        if require_type:
            groups = {b: s for b, s in groups.items() if b.x_file_type == require_type}
        return groups

    def _batch_label(self):
        return (self.batch_id.name or '').replace(' ', '_').replace('/', '-')

    def _bank_label(self, bank):
        return (bank.acc_number or bank.bank_id.name or str(bank.id)
                ).replace(' ', '_').replace('/', '-')

    def _bundle_and_download(self, files):
        if not files:
            raise UserError(_('No files were generated.'))
        batch = self.batch_id
        if len(files) == 1:
            fname, data = files[0]
            mimetype = (
                'text/plain' if fname.endswith('.txt')
                else 'application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet'
            )
            att = self.env['ir.attachment'].create({
                'name': fname, 'type': 'binary',
                'datas': base64.b64encode(data),
                'mimetype': mimetype,
                'res_model': batch._name, 'res_id': batch.id,
            })
        else:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname, data in files:
                    zf.writestr(fname, data)
            att = self.env['ir.attachment'].create({
                'name': 'CommissionsBank_%s.zip' % self._batch_label(),
                'type': 'binary',
                'datas': base64.b64encode(buf.getvalue()),
                'mimetype': 'application/zip',
                'res_model': batch._name, 'res_id': batch.id,
            })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % att.id,
            'target': 'new',
        }

    # --- Excel generation --------------------------------------------------

    def _make_comm_summary_excel(self, wb, sheets):
        """Fill a summary worksheet (mirrors payroll summary but for commissions)."""
        ws = wb.active
        ws.title = 'Commission Summary'
        thin = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('thin'), bottom=Side('thin'),
        )
        bold = Font(bold=True, size=11)
        hdr_fill = PatternFill('solid', fgColor='D9E1F2')
        headers = [
            'Employee', 'SSN', 'Department',
            'Lines Total', 'Driver Commission', 'Gross Total',
            'Loans Deduction', 'Bank Transfer Amount',
            'Bank Account', 'Bank Name',
        ]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = bold
            c.fill = hdr_fill
            c.border = thin
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, sheet in enumerate(
                sheets.sorted(lambda s: s.employee_id.name or ''), 2):
            emp = sheet.employee_id.sudo()
            bank = getattr(emp, 'x_salary_bank_account_id', False)
            row = [
                emp.name or '',
                emp.identification_id or '',
                emp.department_id.name if emp.department_id else '',
                sheet.lines_subtotal,
                sheet.driver_commission_amount,
                sheet.total,
                sheet.x_loans_amount_locked,
                sheet.total_payable,
                bank.acc_number if bank else '',
                bank.bank_id.name if bank and bank.bank_id else '',
            ]
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = thin

    def _make_wps_excel(self, bank, sheets):
        if not openpyxl:
            raise UserError(_('openpyxl is required for Excel export.'))
        wb = openpyxl.Workbook()
        self._make_comm_summary_excel(wb, sheets)

        ws = wb.create_sheet('WPS Bank File')
        thin = Border(left=Side('thin'), right=Side('thin'),
                      top=Side('thin'), bottom=Side('thin'))
        bold = Font(bold=True, size=11)
        hdr_fill = PatternFill('solid', fgColor='C6EFCE')

        cic = bank.x_wps_cic_number or ''
        debit = bank.x_wps_debit_account or ''
        mol = bank.x_wps_mol_id or ''

        ws.cell(1, 1, 'CIC').font = bold
        ws.cell(1, 2, cic)
        ws.cell(2, 1, 'Debit Account:').font = bold
        ws.cell(2, 2, debit)
        ws.cell(3, 1, 'MOL ID').font = bold
        ws.cell(3, 2, mol)
        ws.cell(4, 1, 'Payment Purpose').font = bold
        ws.cell(4, 2, 'Commissions')

        en_headers = [
            'Bank Name', 'Account Number', 'Employee Name', 'Employee Number',
            'National ID', 'Amount (SAR)', 'Basic', 'HRA',
            'Other Earnings', 'Deductions', 'Department',
        ]
        for ci, h in enumerate(en_headers, 1):
            c = ws.cell(6, ci, h)
            c.font = bold
            c.fill = hdr_fill
            c.border = thin
            c.alignment = Alignment(horizontal='center')

        ri = 7
        for sheet in sheets.sorted(lambda s: s.employee_id.name or ''):
            amt = sheet.total_payable
            if not amt:
                continue
            emp = sheet.employee_id.sudo()
            emp_bank = getattr(emp, 'x_salary_bank_account_id', False)
            row = [
                emp_bank.bank_id.name if emp_bank and emp_bank.bank_id else '',
                emp_bank.acc_number if emp_bank else '',
                emp.name or '',
                emp.barcode or '',
                emp.identification_id or '',
                amt,
                sheet.wage,
                0.0,  # HRA not applicable in commissions
                sheet.lines_subtotal + sheet.driver_commission_amount,
                sheet.x_loans_amount_locked,
                emp.department_id.name if emp.department_id else '',
            ]
            for ci, v in enumerate(row, 1):
                ws.cell(ri, ci, v).border = thin
            ri += 1

        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # --- TXT generation (Kawthar format) -----------------------------------

    def _make_kawthar_txt(self, bank, sheets):
        """Generate Kawthar fixed-width 194-char TXT file for commissions."""
        lines = []
        op = (self.operation_code or '2')
        vd = (self.value_date or fields.Date.context_today(self))
        vd_str = vd.strftime('%Y%m%d') if hasattr(vd, 'strftime') else str(vd).replace('-', '')

        for sheet in sheets.sorted(lambda s: s.employee_id.name or ''):
            amt_halala = int(round(sheet.total_payable * 100))
            if amt_halala <= 0:
                continue
            emp = sheet.employee_id.sudo()
            emp_bank = getattr(emp, 'x_salary_bank_account_id', False)
            basic_halala = int(round((sheet.wage or 0.0) * 100))

            barcode = (emp.barcode or '').ljust(12)[:12]
            cic = (bank.x_wps_cic_number or '').ljust(10)[:10]
            card_no = (emp_bank.acc_number if emp_bank else '').ljust(14)[:14]
            emp_name = (emp.name or '').ljust(50)[:50]
            nat_id = (emp.identification_id or '').ljust(10)[:10]
            net_str = str(amt_halala).zfill(15)
            basic_str = str(basic_halala).zfill(12)
            housing_str = '0' * 12
            other_str = str(int(round(
                (sheet.lines_subtotal + sheet.driver_commission_amount) * 100
            ))).zfill(12)
            ded_str = str(int(round(sheet.x_loans_amount_locked * 100))).zfill(12)

            row = (
                barcode + cic + card_no + emp_name + nat_id
                + net_str + vd_str + op + '0' * 6 + ' ' * 20
                + basic_str + housing_str + other_str + ded_str
            )
            assert len(row) == 194, f"Row length {len(row)} != 194"
            lines.append(row)

        return '\n'.join(lines)

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------

    def action_export(self):
        self.ensure_one()
        mode = self.export_mode
        handlers = {
            'all_excel':     self._export_all_excel,
            'all_txt':       self._export_all_txt,
            'specific_excel': self._export_specific_excel,
            'specific_txt':  self._export_specific_txt,
        }
        return handlers[mode]()

    def _export_all_excel(self):
        groups = self._group_and_validate(require_type=None)
        files = []
        bl = self._batch_label()
        for bank, sheets in groups.items():
            label = self._bank_label(bank)
            if bank.x_file_type in ('wps', 'kawthar'):
                data = self._make_wps_excel(bank, sheets)
                files.append(('Commissions_%s_%s.xlsx' % (bl, label), data))
        if not files:
            raise UserError(_('No Excel files could be generated.'))
        return self._bundle_and_download(files)

    def _export_all_txt(self):
        groups = self._group_and_validate(require_type='kawthar')
        files = []
        bl = self._batch_label()
        vd = (self.value_date or fields.Date.context_today(self)
              ).strftime('%Y%m%d')
        for bank, sheets in groups.items():
            label = self._bank_label(bank)
            data = self._make_kawthar_txt(bank, sheets).encode('utf-8')
            files.append(('Commissions_%s_%s_%s.txt' % (bl, label, vd), data))
        if not files:
            raise UserError(_('No text files could be generated.'))
        return self._bundle_and_download(files)

    def _export_specific_excel(self):
        if not self.bank_account_id:
            raise UserError(_('Please select a bank account.'))
        bank = self.bank_account_id
        groups = self._group_and_validate()
        sheets = groups.get(bank)
        if not sheets:
            raise UserError(_('No sheets are assigned to the selected bank.'))
        bl = self._batch_label()
        label = self._bank_label(bank)
        data = self._make_wps_excel(bank, sheets)
        return self._bundle_and_download(
            [('Commissions_%s_%s.xlsx' % (bl, label), data)])

    def _export_specific_txt(self):
        if not self.bank_account_id:
            raise UserError(_('Please select a bank account.'))
        if not self.value_date:
            raise UserError(_('Please set a value date for the text file.'))
        bank = self.bank_account_id
        groups = self._group_and_validate()
        sheets = groups.get(bank)
        if not sheets:
            raise UserError(_('No sheets are assigned to the selected bank.'))
        bl = self._batch_label()
        label = self._bank_label(bank)
        vd = self.value_date.strftime('%Y%m%d')
        data = self._make_kawthar_txt(bank, sheets).encode('utf-8')
        return self._bundle_and_download(
            [('Commissions_%s_%s_%s.txt' % (bl, label, vd), data)])

