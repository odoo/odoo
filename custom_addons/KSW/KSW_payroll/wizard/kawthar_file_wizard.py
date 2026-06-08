import base64
import io

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    openpyxl = None

# Operating codes from the Al Rajhi Kawthar/Itqan template
OPERATION_CODES = [
    ('1', 'Create New Card - أنشاء بطاقة جديدة'),
    ('2', 'Load Funds - تحميل رصيد'),
    ('3', 'Close Card - أغلاق البطاقة'),
    ('4', 'Activate Card - تنشيط البطاقة'),
    ('5', 'Mark Card as Lost - ألاشارة على البطاقة كضائعة'),
    ('6', 'Mark Card as Stolen - ألاشارة على البطاقة كمسروقة'),
    ('7', 'Re-Issue Card - أعادة اصدار البطاقة'),
    ('8', 'Partial Refund - أستراد جزء من المبلغ'),
    ('9', 'Full Refund - أستراد كامل المبلغ'),
    ('A', 'Update Employee ID - تحديث الرقم الوظيفي'),
    ('B', 'Update National ID - تحديث الهوية'),
    ('C', 'Update Employee Name - تحديث اسم الموظف'),
    ('D', 'Change Card Details & Special Close - تحديث المعلومات كاملة'),
    ('S', 'Special Close - إقفال بطاقة وإصدار بطاقة'),
    ('P', 'Create New PIN - تجديد الرقم السري'),
]

# Arabic labels for operation-type columns (O–AD) in the template header
OPERATION_LABELS_AR = [
    'أنشاء بطاقة جديدة',
    'تحميل رصيد ',
    'أغلاق البطاقة',
    'تنشيط البطاقة',
    'ألاشارة على البطاقة كضائعة',
    'ألاشارة على البطاقة كمسروقة',
    'أعادة اصدار البطاقة',
    'أستراد جزء من المبلغ',
    'أستراد كامل المبلغ',
    'تحديث الرقم الوظيفي',
    'تحديث الرقم الهوية',
    'تحديث أسم الموظف',
    'طلب تغيير تفاصيل البطاقة',
    'تجديد البطاقة سوف تنتهي',
    'طلب رقم سري جديد',
    'إغلاق و إنشاء خاص',
]

OPERATION_LABELS_EN = [
    'Create New Card',
    'Load Funds',
    'Close Card',
    'Activate Card',
    'Mark Card as Lost',
    'Mark Card as Stolen',
    'Re-Issue Card',
    'Partial Refund',
    'Full Refund',
    'Update Employee Id',
    'Update National Id',
    'Update Employee Name',
    'Change Card Details and Speical Close Card',
    'Renew of expiring Card',
    'Create New Pin',
    'Special Close',
]

# Map operation_code selection value → Arabic label for the data column
OPERATION_CODE_LABEL = {
    '1': 'أنشاء بطاقة جديدة',
    '2': 'تحميل رصيد ',
    '3': 'أغلاق البطاقة',
    '4': 'تنشيط البطاقة',
    '5': 'ألاشارة على البطاقة كضائعة',
    '6': 'ألاشارة على البطاقة كمسروقة',
    '7': 'أعادة اصدار البطاقة',
    '8': 'أستراد جزء من المبلغ',
    '9': 'أستراد كامل المبلغ',
    'A': 'تحديث الرقم الوظيفي',
    'B': 'تحديث الهوية',
    'C': 'تحديث أسم الموظف',
    'D': 'طلب تغيير تفاصيل البطاقة',
    'S': 'إقفال بطاقة وإصدار بطاقة',
    'P': 'طلب رقم سري جديد',
}


class KawtharFileWizard(models.TransientModel):
    _name = 'ksw.kawthar.file.wizard'
    _description = 'Kawthar (Itqan) Payroll Card File Generator'

    payslip_run_id = fields.Many2one(
        'hr.payslip.run', string='Payslip Batch',
        required=True, readonly=True,
    )
    operation_code = fields.Selection(
        OPERATION_CODES,
        string='Operation',
        required=True,
        default='2',
        help='Type of payroll-card operation to include in the file.',
    )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_line_total(self, slip, code):
        """Return the total of a salary-rule line by code, or 0."""
        line = slip.line_ids.filtered(lambda l: l.code == code)
        return line[:1].total if line else 0.0

    def _sar_to_halalas(self, amount):
        """Convert SAR amount to halalas (×100)."""
        return int(round(amount * 100))

    def _pz(self, number, length):
        """Zero-pad integer number to length digits."""
        return str(int(number)).zfill(length)[:length]

    def _pr(self, text, length):
        """Left-justify text, pad/truncate to length with spaces."""
        return str(text or '')[:length].ljust(length)

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------

    def action_generate(self):
        """Generate and download the Kawthar Excel file."""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('The openpyxl library is required.'))

        batch = self.payslip_run_id
        if not batch.slip_ids:
            raise UserError(_('No payslips in this batch to export.'))

        # Filter to employees with a bank account and positive NET
        valid_slips = batch._sorted_export_slips(batch.slip_ids.filtered(
            lambda s: s.employee_id.sudo().primary_bank_account_id
            and self._get_line_total(s, 'NET') > 0
        ))

        if not valid_slips:
            raise UserError(_(
                'No payslips with a positive NET salary and a '
                'bank account on the employee.'
            ))

        wb = openpyxl.Workbook()
        self._fill_kawthar_sheet(wb, valid_slips)

        # Remove default empty sheet if still present
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        output = io.BytesIO()
        wb.save(output)
        file_data = base64.b64encode(output.getvalue())

        filename = 'Kawthar_%s.xls' % (
            batch.name.replace(' ', '_').replace('/', '-'),
        )
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.sheet',
            'res_model': batch._name,
            'res_id': batch.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }

    # ------------------------------------------------------------------
    # Sheet builder
    # ------------------------------------------------------------------

    def _fill_kawthar_sheet(self, wb, slips):
        """Build the PREFORMAT PAYMENTS sheet in the Al Rajhi Kawthar
        template format."""
        ws = wb.create_sheet('PREFORMAT PAYMENTS')

        bold = Font(bold=True, size=11)
        title_font = Font(bold=True, size=12)
        thin = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('thin'), bottom=Side('thin'),
        )
        hdr_fill = PatternFill('solid', fgColor='D9E1F2')

        # ── Row 1: Title + operation-type column labels (EN) ──
        ws.cell(1, 1, 'Payroll Cards File Upload').font = title_font
        for ci, label in enumerate(OPERATION_LABELS_EN):
            ws.cell(1, 16 + ci, label).font = bold  # columns P onwards (16-based)

        # ── Row 2: Notes ──
        ws.cell(2, 1,
                'Notes : Do Not change the order of the columns or '
                'delete it, all detials must be in English.').font = Font(
            italic=True, size=10,
        )

        # ── Row 3: English headers ──
        en_headers = [
            'Employee ID (12N)',
            'Payroll Card No.  (19N)',
            'Employee Name (50A)',
            'National ID Number (10N)',
            'Amount (15N)',
            'Operating Code (1X)',
            'Department ID(5N)',
            'Basic Salary(12)',
            'Hoousing allowance(12N)',
            'Other Allowance(12N)',
            'Deductions(12N)',
            'Mobile Number(10N)',
            'Bio Pin(1N)',
            'User Field(10A)',
        ]
        for ci, h in enumerate(en_headers, 1):
            c = ws.cell(3, ci, h)
            c.font = bold
            c.fill = hdr_fill
            c.border = thin
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        # Operation-type column Arabic labels in row 3
        for ci, label in enumerate(OPERATION_LABELS_AR):
            c = ws.cell(3, 16 + ci, label)
            c.font = bold
            c.fill = hdr_fill
            c.border = thin
            c.alignment = Alignment(horizontal='center')

        # ── Row 4: Arabic headers (main columns only) ──
        ar_headers = [
            'الرقم الوظيفي',
            'رقم بطاقة اتقان',
            'اسم الموظف',
            'رقم الموظف',
            'المبلغ',
            'نوع العملية',
        ]
        for ci, h in enumerate(ar_headers, 1):
            c = ws.cell(4, ci, h)
            c.font = Font(bold=True, size=10)
            c.border = thin
            c.alignment = Alignment(horizontal='center')

        # ── Data rows (row 5+) ──
        op_label = OPERATION_CODE_LABEL.get(self.operation_code, '')
        seq = 0
        for slip in slips:
            emp = slip.employee_id
            net = self._get_line_total(slip, 'NET')
            if not net:
                continue

            seq += 1
            bank = emp.sudo().primary_bank_account_id
            basic = self._get_line_total(slip, 'BASIC')
            hra = self._get_line_total(slip, 'HRA')
            gross = self._get_line_total(slip, 'GROSS')
            other = max(gross - basic - hra, 0.0) if gross else 0.0

            total_ded = abs(sum(
                l.total for l in slip.line_ids
                if l.category_id.code == 'DED'
            )) or 0.0

            row = [
                seq,                                                   # A: Employee ID (sequential)
                bank.acc_number or '' if bank else '',                 # B: Payroll Card No.
                (emp.name or '').upper(),                              # C: Employee Name
                emp.ssnid or emp.identification_id or '',              # D: National ID
                int(round(net)),                                       # E: Amount (whole SAR)
                op_label,                                              # F: Operating Code
                '',                                                    # G: Department ID
                int(round(basic)) if basic else 0,                     # H: Basic Salary
                int(round(hra)) if hra else 0,                         # I: Housing Allowance
                int(round(other)) if other else 0,                     # J: Other Allowance
                int(round(total_ded)) if total_ded else 0,             # K: Deductions
                '',                                                    # L: Mobile Number
                '',                                                    # M: Bio Pin
                '',                                                    # N: User Field
            ]
            ri = 4 + seq  # row 5, 6, 7, …
            for ci, v in enumerate(row, 1):
                c = ws.cell(ri, ci, v)
                c.border = thin

        # ── Auto-width columns ──
        for ci in range(1, len(en_headers) + 1):
            letter = openpyxl.utils.get_column_letter(ci)
            mx = max(
                len(str(ws.cell(row=r, column=ci).value or ''))
                for r in range(1, max(ws.max_row + 1, 2))
            )
            ws.column_dimensions[letter].width = min(mx + 3, 40)

    # ------------------------------------------------------------------
    # Text-file builder (Al Rajhi Kawthar/Itqan import format)
    # ------------------------------------------------------------------

    def _build_kawthar_text(self, bank_account, slips, value_date):
        """Build the fixed-width Kawthar text file content.

        Each line is 194 characters, no header row.
        Format per line:
          Employee barcode (12N) | CIC (10N) | Card number (14N) |
          Employee name (50A) | National ID (10N) | Net in halalas (15N) |
          Value date YYYYMMDD (8) | Operation code (1X) |
          Zeros filler (6) | Spaces filler (20) |
          Basic in halalas (12N) | Housing in halalas (12N) |
          Other allowance in halalas (12N) | Deductions in halalas (12N)
        """
        valid = self.payslip_run_id._sorted_export_slips(slips.filtered(
            lambda s: s.employee_id.sudo().primary_bank_account_id
            and self._get_line_total(s, 'NET') > 0
        ))

        if not valid:
            from odoo.exceptions import UserError
            raise UserError(_('No payslips with a positive NET and a bank account.'))

        cic = (bank_account.x_wps_cic_number or '').replace(' ', '')
        op = self.operation_code or '2'
        vdate = value_date.strftime('%Y%m%d')

        lines = []
        for slip in valid:
            emp = slip.employee_id
            bank = emp.sudo().primary_bank_account_id

            emp_ref = emp.barcode or '0'
            card_no = (bank.acc_number or '').replace(' ', '')
            emp_name = (emp.name or '')
            emp_id = emp.ssnid or emp.identification_id or '0'

            net = self._get_line_total(slip, 'NET')
            basic = self._get_line_total(slip, 'BASIC')
            hra = self._get_line_total(slip, 'HRA')
            gross = self._get_line_total(slip, 'GROSS')
            other = max(gross - basic - hra, 0.0) if gross else 0.0
            total_ded = abs(sum(
                l.total for l in slip.line_ids
                if l.category_id.code == 'DED'
            )) or 0.0

            line = ''.join([
                self._pz(int(emp_ref) if emp_ref.isdigit() else 0, 12),
                self._pz(int(cic) if cic.isdigit() else 0, 10),
                self._pr(card_no, 14),
                self._pr(emp_name, 50),
                self._pz(int(emp_id) if emp_id.isdigit() else 0, 10),
                self._pz(self._sar_to_halalas(net), 15),
                vdate,
                str(op)[:1],
                '0' * 6,
                ' ' * 20,
                self._pz(self._sar_to_halalas(basic), 12),
                self._pz(self._sar_to_halalas(hra), 12),
                self._pz(self._sar_to_halalas(other), 12),
                self._pz(self._sar_to_halalas(total_ded), 12),
            ])
            lines.append(line)

        return '\n'.join(lines) + '\n'

