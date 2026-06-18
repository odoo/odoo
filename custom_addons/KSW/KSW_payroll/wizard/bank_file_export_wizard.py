import base64
import io
import zipfile

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

EXPORT_MODES = [
    ('all_excel', 'All banks – Excel files'),
    ('all_txt', 'All banks – Text files'),
    ('specific_excel', 'Specific bank – Excel'),
    ('specific_txt', 'Specific bank – Text file'),
]


class BankFileExportWizard(models.TransientModel):
    _name = 'ksw.bank.file.export.wizard'
    _description = 'Unified Bank File Export Wizard'

    payslip_run_id = fields.Many2one(
        'hr.payslip.run', string='Payslip Batch',
        required=True, readonly=True,
    )
    export_mode = fields.Selection(
        EXPORT_MODES,
        string='Export Mode',
        required=True,
        default='all_excel',
    )
    bank_account_id = fields.Many2one(
        'res.partner.bank',
        string='Bank Account',
        domain="[('x_file_type', '!=', False),"
               " ('partner_id', '=', company_partner_id)]",
    )
    company_partner_id = fields.Many2one(
        'res.partner',
        compute='_compute_company_partner_id',
    )

    @api.depends('payslip_run_id')
    def _compute_company_partner_id(self):
        for rec in self:
            rec.company_partner_id = rec.env.company.partner_id
    value_date = fields.Date(
        string='Value Date',
        default=fields.Date.context_today,
        help='Payment value date for the WPS text file header.',
    )

    # Import operation codes from the Kawthar wizard module
    operation_code = fields.Selection(
        selection='_get_operation_codes',
        string='Kawthar Operation',
        default='2',
        help='Operation type for Kawthar payroll card files.',
    )

    @api.model
    def _get_operation_codes(self):
        from .kawthar_file_wizard import OPERATION_CODES
        return OPERATION_CODES

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _group_and_validate(self, require_type=None):
        """Group slips by bank and validate.

        :param require_type: if set, filter to banks with this x_file_type
        :returns: dict {res.partner.bank: hr.payslip recordset}
        """
        batch = self.payslip_run_id
        if not batch.slip_ids:
            raise UserError(_('No payslips in this batch to export.'))

        groups = batch._group_slips_by_bank_account()

        # Check for employees without any bank
        no_bank = groups.pop(self.env['res.partner.bank'], None)
        if no_bank:
            names = ', '.join(no_bank.mapped('employee_id.name'))
            raise UserError(_(
                'The following employees have no Salary Paying Bank Account '
                'set (and no fallback on the batch):\n%s', names,
            ))

        # Validate every bank account has x_file_type configured
        no_type = [
            b for b in groups
            if not b.x_file_type
        ]
        if no_type:
            names = ', '.join(b.acc_number or b.bank_id.name or str(b.id)
                              for b in no_type)
            raise UserError(_(
                'The following bank accounts have no Payroll File Type set. '
                'Please configure it on each bank account:\n%s', names,
            ))

        # Filter by type if requested
        if require_type:
            groups = {b: s for b, s in groups.items()
                      if b.x_file_type == require_type}

        return groups

    def _make_wps_excel(self, bank, slips):
        """Generate a WPS Excel workbook (summary + bank sheet) as bytes."""
        if not openpyxl:
            raise UserError(_('The openpyxl library is required.'))
        batch = self.payslip_run_id
        wb = openpyxl.Workbook()
        batch._fill_payroll_summary_sheet(wb, slips)
        batch._fill_wps_sheet(wb, bank, slips)
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _make_wps_txt(self, bank, slips):
        """Generate a WPS text file as bytes."""
        wps_wiz = self.env['ksw.wps.file.wizard'].new({
            'payslip_run_id': self.payslip_run_id.id,
            'value_date': self.value_date,
        })
        content = wps_wiz._build_wps_text(bank, slips)
        return content.encode('utf-8')

    def _make_kawthar_excel(self, slips):
        """Generate a Kawthar Excel workbook as bytes."""
        if not openpyxl:
            raise UserError(_('The openpyxl library is required.'))
        kaw_wiz = self.env['ksw.kawthar.file.wizard'].new({
            'payslip_run_id': self.payslip_run_id.id,
            'operation_code': self.operation_code or '2',
        })
        batch = self.payslip_run_id
        wb = openpyxl.Workbook()
        # Filter to employees with bank account and positive NET
        valid_slips = batch._sorted_export_slips(slips.filtered(
            lambda s: s.employee_id.sudo().primary_bank_account_id
            and kaw_wiz._get_line_total(s, 'NET') > 0
        ))
        if not valid_slips:
            return None
        batch._fill_payroll_summary_sheet(wb, valid_slips)
        kaw_wiz._fill_kawthar_sheet(wb, valid_slips)
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _make_kawthar_txt(self, bank, slips):
        """Generate a Kawthar text file as bytes."""
        kaw_wiz = self.env['ksw.kawthar.file.wizard'].new({
            'payslip_run_id': self.payslip_run_id.id,
            'operation_code': self.operation_code or '2',
        })
        content = kaw_wiz._build_kawthar_text(bank, slips, self.value_date)
        return content.encode('utf-8')

    def _bank_label(self, bank):
        """Short label for a bank account, safe for filenames."""
        label = (bank.acc_number or bank.bank_id.name or str(bank.id))
        return label.replace(' ', '_').replace('/', '-')

    def _batch_label(self):
        return self.payslip_run_id.name.replace(' ', '_').replace('/', '-')

    def _bundle_and_download(self, files):
        """Create attachment(s) and return download action.

        :param files: list of (filename, bytes) tuples
        """
        batch = self.payslip_run_id
        if not files:
            raise UserError(_('No files were generated.'))

        if len(files) == 1:
            fname, data = files[0]
            mimetype = (
                'text/plain' if fname.endswith('.txt')
                else 'application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet'
            )
            att = self.env['ir.attachment'].create({
                'name': fname,
                'type': 'binary',
                'datas': base64.b64encode(data),
                'mimetype': mimetype,
                'res_model': batch._name,
                'res_id': batch.id,
            })
        else:
            # Bundle into a zip
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, 'w',
                                 zipfile.ZIP_DEFLATED) as zf:
                for fname, data in files:
                    zf.writestr(fname, data)
            zip_name = 'BankFiles_%s.zip' % self._batch_label()
            att = self.env['ir.attachment'].create({
                'name': zip_name,
                'type': 'binary',
                'datas': base64.b64encode(zip_buf.getvalue()),
                'mimetype': 'application/zip',
                'res_model': batch._name,
                'res_id': batch.id,
            })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % att.id,
            'target': 'new',
        }

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------

    def action_export(self):
        """Generate and download bank files based on the selected mode."""
        self.ensure_one()
        mode = self.export_mode
        handler = {
            'all_excel': self._export_all_excel,
            'all_txt': self._export_all_txt,
            'specific_excel': self._export_specific_excel,
            'specific_txt': self._export_specific_txt,
        }
        return handler[mode]()

    # -- Mode handlers -----------------------------------------------------

    def _export_all_excel(self):
        groups = self._group_and_validate()
        files = []
        bl = self._batch_label()
        for bank, slips in groups.items():
            label = self._bank_label(bank)
            if bank.x_file_type == 'wps':
                data = self._make_wps_excel(bank, slips)
                files.append(('WPS_%s_%s.xlsx' % (bl, label), data))
            elif bank.x_file_type == 'kawthar':
                data = self._make_kawthar_excel(slips)
                if data:
                    files.append(('Kawthar_%s_%s.xlsx' % (bl, label), data))
        if not files:
            raise UserError(_('No Excel files could be generated.'))
        return self._bundle_and_download(files)

    def _export_all_txt(self):
        groups = self._group_and_validate()
        if not self.value_date:
            raise UserError(_('Please set a Value Date for the text file.'))
        files = []
        bl = self._batch_label()
        vd = self.value_date.strftime('%Y%m%d')
        for bank, slips in groups.items():
            label = self._bank_label(bank)
            if bank.x_file_type == 'wps':
                data = self._make_wps_txt(bank, slips)
                files.append(('WPS_%s_%s_%s.txt' % (bl, label, vd), data))
            elif bank.x_file_type == 'kawthar':
                data = self._make_kawthar_txt(bank, slips)
                files.append(('Kawthar_%s_%s_%s.txt' % (bl, label, vd), data))
        if not files:
            raise UserError(_('No text files could be generated.'))
        return self._bundle_and_download(files)

    def _export_specific_excel(self):
        if not self.bank_account_id:
            raise UserError(_('Please select a bank account.'))
        bank = self.bank_account_id
        if not bank.x_file_type:
            raise UserError(_(
                'The selected bank account has no Payroll File Type set.'))
        groups = self._group_and_validate()
        slips = groups.get(bank)
        if not slips:
            raise UserError(_(
                'No payslips are assigned to the selected bank account.'))
        bl = self._batch_label()
        label = self._bank_label(bank)
        if bank.x_file_type == 'wps':
            data = self._make_wps_excel(bank, slips)
            fname = 'WPS_%s_%s.xlsx' % (bl, label)
        elif bank.x_file_type == 'kawthar':
            data = self._make_kawthar_excel(slips)
            if not data:
                raise UserError(_(
                    'No payslips with a positive NET and a bank account '
                    'on the employee for the Kawthar file.'))
            fname = 'Kawthar_%s_%s.xlsx' % (bl, label)
        else:
            raise UserError(_('Unsupported file type: %s', bank.x_file_type))
        return self._bundle_and_download([(fname, data)])

    def _export_specific_txt(self):
        if not self.bank_account_id:
            raise UserError(_('Please select a bank account.'))
        bank = self.bank_account_id
        if not bank.x_file_type:
            raise UserError(_(
                'The selected bank account has no Payroll File Type set.'))
        if not self.value_date:
            raise UserError(_('Please set a Value Date for the text file.'))
        groups = self._group_and_validate()
        slips = groups.get(bank)
        if not slips:
            raise UserError(_(
                'No payslips are assigned to the selected bank account.'))
        bl = self._batch_label()
        label = self._bank_label(bank)
        vd = self.value_date.strftime('%Y%m%d')
        if bank.x_file_type == 'wps':
            data = self._make_wps_txt(bank, slips)
            fname = 'WPS_%s_%s_%s.txt' % (bl, label, vd)
        elif bank.x_file_type == 'kawthar':
            data = self._make_kawthar_txt(bank, slips)
            fname = 'Kawthar_%s_%s_%s.txt' % (bl, label, vd)
        else:
            raise UserError(_('Unsupported file type: %s', bank.x_file_type))
        return self._bundle_and_download([(fname, data)])
