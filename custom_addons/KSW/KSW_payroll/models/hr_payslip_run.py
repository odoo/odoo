from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    openpyxl = None


class KswPayslipRunSkipLine(models.Model):
    """Records employees that were skipped during batch payslip generation,
    along with the reason they were excluded."""
    _name = 'ksw.payslip.run.skip.line'
    _description = 'Payslip Batch — Skipped Employee Log'
    _order = 'employee_id'

    run_id = fields.Many2one(
        'hr.payslip.run', string='Payslip Batch',
        required=True, ondelete='cascade', index=True,
    )
    employee_id = fields.Many2one(
        'hr.employee', string='Employee',
        required=True, ondelete='cascade',
    )
    reason = fields.Char(string='Reason', required=True)


class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'

    x_salary_bank_account_id = fields.Many2one(
        'res.partner.bank',
        string='Salary Paying Bank Account',
        help='Default company bank account for WPS export. '
             'Employees with their own Salary Paying Bank Account '
             'will override this.',
    )

    x_skip_line_ids = fields.One2many(
        'ksw.payslip.run.skip.line', 'run_id',
        string='Skipped Employees',
        help='Employees that were automatically skipped during payslip '
             'generation and the reason they were excluded.',
    )

    def action_clear_skip_log(self):
        """Remove all skipped-employee log entries for this batch."""
        self.ensure_one()
        self.x_skip_line_ids.unlink()
        return True


    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_line_total(self, slip, code):
        """Return the total of a salary rule line by code, or 0."""
        line = slip.line_ids.filtered(lambda l: l.code == code)
        return line[:1].total if line else 0.0

    def _get_wd_amount(self, slip, code):
        """Return the amount of a worked-day line by code, or 0."""
        wd = slip.worked_days_line_ids.filtered(lambda w: w.code == code)
        return wd[:1].amount if wd else 0.0

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------

    def action_open_export_wizard(self):
        """Open the unified bank file export wizard."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Export Bank File'),
            'res_model': 'ksw.bank.file.export.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_payslip_run_id': self.id,
            },
        }

    def action_open_batch_payslips(self):
        """Open batch payslip lines in the standard searchable payslip view."""
        self.ensure_one()
        action = self.env['ir.actions.actions']._for_xml_id(
            'om_hr_payroll.action_view_hr_payslip_form'
        )
        action.update({
            'name': _('Payslips - %s') % (self.name or _('Batch')),
            'domain': [('payslip_run_id', '=', self.id)],
            'context': {
                'default_payslip_run_id': self.id,
                'search_default_payslip_run_id': self.id,
            },
        })
        return action

    def _group_slips_by_bank_account(self):
        """Group payslips by the paying bank account

        Resolution order for each slip:
        1. Employee's ``x_salary_bank_account_id``
        2. Batch-level ``x_salary_bank_account_id`` (fallback)

        Returns a dict ``{res.partner.bank recordset: slip recordsets}``.
        Slips with **no** resolved bank account are collected under an
        empty recordset key so callers can report them.
        """
        groups = {}
        bank_model = self.env['res.partner.bank']
        for slip in self.slip_ids:
            bank = (
                slip.employee_id.sudo().x_salary_bank_account_id
                or self.x_salary_bank_account_id
                or bank_model
            )
            groups.setdefault(bank, self.env['hr.payslip'])
            groups[bank] |= slip
        return groups

    def _sorted_export_slips(self, slips):
        """Sort slips by configured employee export order, then name.

        ``x_payslip_export_order`` values <= 0 are treated as unset and sent
        to the end so explicitly configured employees always come first.
        """
        return slips.sorted(
            lambda s: (
                not bool(s.employee_id.sudo().x_payslip_export_order > 0),
                s.employee_id.sudo().x_payslip_export_order
                if s.employee_id.sudo().x_payslip_export_order > 0 else 0,
                s.employee_id.name or '',
                s.employee_id.id,
            )
        )

    # ------------------------------------------------------------------
    # Sheet 1 — Internal payroll summary
    # ------------------------------------------------------------------

    def _fill_payroll_summary_sheet(self, wb, slips=None):
        if slips is None:
            slips = self.slip_ids
        ws = wb.active
        ws.title = 'Payroll Summary'

        headers = [
            'Employee', 'SSN No', 'Date From', 'Date To', 'Department',
            'Basic Salary', 'House Rent Allowance', 'Other Allowance',
            'Gross', 'Absence Deduction', 'Attendance Deductions',
            'Missed Days (ATT SHEET)', 'Social Insurance', 'Loan',
            'Net Salary', 'Bank Account Number', 'Bank Name',
        ]

        hdr_font = Font(bold=True, size=11)
        hdr_fill = PatternFill('solid', fgColor='D9E1F2')
        hdr_align = Alignment(horizontal='center', wrap_text=True)
        thin = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('thin'), bottom=Side('thin'),
        )

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = thin

        for ri, slip in enumerate(self._sorted_export_slips(slips), 2):
            emp = slip.employee_id
            bank = emp.sudo().primary_bank_account_id
            is_sheet = emp.sudo().x_is_attendance_sheet

            basic = self._get_line_total(slip, 'BASIC')
            hra = self._get_line_total(slip, 'HRA')
            gross = self._get_line_total(slip, 'GROSS')
            gosi = self._get_line_total(slip, 'GOSI')
            net = self._get_line_total(slip, 'NET')
            attded = self._get_line_total(slip, 'ATTDED')

            # Other allowance = GROSS - BASIC - HRA
            other_alw = gross - basic - hra if gross else 0.0

            # Deductions breakdown
            abs_ded = 0.0
            att_ded = 0.0
            sheet_ded = 0.0
            if is_sheet:
                # Sheet employee: deduction goes to column L
                sheet_ded = -self._get_wd_amount(slip, 'ATT_DED')
            else:
                # Biometric: absence (J) and late/early (K)
                abs_ded = -self._get_wd_amount(slip, 'ATT_ABS')
                att_ded = -(
                    self._get_wd_amount(slip, 'ATT_LATE')
                    + self._get_wd_amount(slip, 'ATT_EARLY')
                )

            # Loan = total DED minus ATTDED and GOSI
            total_ded_cat = sum(
                l.total for l in slip.line_ids
                if l.category_id.code == 'DED'
            )
            loan = total_ded_cat - attded - gosi

            row = [
                emp.name or '',
                emp.ssnid or emp.identification_id or '',
                slip.date_from,
                slip.date_to,
                emp.department_id.name if emp.department_id else '',
                basic or None,
                hra or None,
                other_alw or None,
                gross or None,
                abs_ded or None,
                att_ded or None,
                sheet_ded or None,
                gosi or None,
                loan or None,
                net,
                bank.acc_number if bank else '',
                bank.bank_id.name if bank and bank.bank_id else '',
            ]
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = thin

        # Auto-width columns
        for ci in range(1, len(headers) + 1):
            letter = openpyxl.utils.get_column_letter(ci)
            mx = max(
                len(str(ws.cell(row=r, column=ci).value or ''))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = min(mx + 3, 35)

    # ------------------------------------------------------------------
    # Sheet 2 — WPS bank upload (Al Rajhi format)
    # ------------------------------------------------------------------

    def _fill_wps_sheet(self, wb, bank_account, slips=None, suffix=''):
        if slips is None:
            slips = self.slip_ids
        sheet_name = ('WPS Bank File%s' % suffix)[:31]  # Excel 31-char limit
        ws = wb.create_sheet(sheet_name)

        bold = Font(bold=True, size=11)
        title_font = Font(bold=True, size=12, color='003366')
        thin = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('thin'), bottom=Side('thin'),
        )
        hdr_fill = PatternFill('solid', fgColor='C6EFCE')
        ar_fill = PatternFill('solid', fgColor='E2EFDA')

        # Read WPS header values from the selected bank account
        cic = bank_account.x_wps_cic_number or ''
        debit = bank_account.x_wps_debit_account or ''
        mol = bank_account.x_wps_mol_id or ''

        # ── Row 1 ──
        ws.cell(1, 1, 'CIC - رقم العميل').font = bold
        ws.cell(1, 2, cic)
        ws.merge_cells('C1:D1')
        ws.cell(1, 3,
                'Alrajhi Bank WPS Payroll Payments Upload File'
                ).font = title_font

        # ── Row 2 ──
        ws.cell(2, 1, 'Debit Account:').font = bold
        ws.cell(2, 2, debit)
        ws.merge_cells('C2:D2')
        ws.cell(2, 3,
                'Notes: Template used for upload of WPS Payroll data'
                ).font = Font(bold=True, size=10)
        ws.cell(2, 8, 'Type of Payroll')
        ws.cell(2, 9, 'WPS')

        # ── Row 3 ──
        ws.cell(3, 1, 'MOL ID').font = bold
        ws.cell(3, 2, mol)

        # ── Row 4 ──
        ws.cell(4, 1, 'Payment Purpose').font = bold
        ws.cell(4, 2, 'Payroll')

        # ── Row 5 ──
        ws.cell(5, 1, 'Company Remarks').font = bold
        ws.cell(5, 2, 'Payroll')

        # ── Row 6–7: English / Arabic headers ──
        en_headers = [
            'Bank Name', 'Account Number(34N)', 'Employee Name',
            'Employee Number', 'National ID Number', 'Salary (15N)',
            'Basic Salary', 'Housing Allowance', 'Other Earnings',
            'Deductions', 'Branch Code', 'Branch Name',
            'Employee Remarks', 'Employee Department',
        ]
        ar_headers = [
            'بنك الموظف', 'رقم أيبان الموظف', 'إسم الموظف',
            'الرقم الوظيفي', 'رقم الهوية للموظف', 'إجمالي الراتب',
            'الراتب الأساسي', 'بدل السكن', 'بدل أخرى',
            'الخصومات', 'رمز الفرع', 'اسم الفرع',
            'ملاحظات الموظف', 'قسم الموظف',
        ]

        for ci, (e, a) in enumerate(zip(en_headers, ar_headers), 1):
            c6 = ws.cell(6, ci, e)
            c6.font = bold
            c6.fill = hdr_fill
            c6.border = thin
            c6.alignment = Alignment(horizontal='center')

            c7 = ws.cell(7, ci, a)
            c7.font = Font(bold=True, size=10)
            c7.fill = ar_fill
            c7.border = thin
            c7.alignment = Alignment(horizontal='center')

        # ── Data rows (row 8+), skip employees with net == 0 ──
        ri = 8
        for slip in self._sorted_export_slips(slips):
            net = self._get_line_total(slip, 'NET')
            if not net:
                continue

            emp = slip.employee_id
            bank = emp.sudo().primary_bank_account_id

            basic = self._get_line_total(slip, 'BASIC')
            hra = self._get_line_total(slip, 'HRA')
            gross = self._get_line_total(slip, 'GROSS')
            other_e = gross - basic - hra if gross else 0.0

            total_ded = abs(sum(
                l.total for l in slip.line_ids
                if l.category_id.code == 'DED'
            )) or 0.0

            data = [
                bank.bank_id.name if bank and bank.bank_id else '',
                bank.acc_number if bank else '',
                emp.name or '',
                emp.barcode or '',
                emp.ssnid or emp.identification_id or '',
                net,              # F: Salary (net)
                basic,            # G: Basic
                hra,              # H: Housing
                other_e,          # I: Other earnings
                total_ded,        # J: Deductions
                '',               # K: Branch Code
                '',               # L: Branch Name
                '',               # M: Employee Remarks
                emp.department_id.name if emp.department_id else '',
            ]
            for ci, v in enumerate(data, 1):
                c = ws.cell(ri, ci, v)
                c.border = thin
            ri += 1

        # Auto-width columns
        for ci in range(1, len(en_headers) + 1):
            letter = openpyxl.utils.get_column_letter(ci)
            mx = max(
                len(str(ws.cell(row=r, column=ci).value or ''))
                for r in range(1, max(ws.max_row + 1, 2))
            )
            ws.column_dimensions[letter].width = min(mx + 3, 40)

