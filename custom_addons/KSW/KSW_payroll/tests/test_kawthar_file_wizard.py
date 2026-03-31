# -*- coding: utf-8 -*-
"""Tests for the Kawthar (Itqan) payroll card file wizard.

Covers:
  - Wizard defaults & constraints
  - Sheet structure: title row, notes, English/Arabic headers
  - Data rows: sequential Employee ID, card numbers, amounts in whole SAR
  - Operation code mapping
  - Zero-net slips skipped; employees without bank account skipped
  - Error handling: empty batch, no valid slips
  - action_open_kawthar_wizard returns correct window action
  - Salary breakdown: basic, hra, other, deductions
"""
import base64
import io
from datetime import date

from odoo.exceptions import UserError
from odoo.tests.common import TransactionCase

try:
    import openpyxl
except ImportError:
    openpyxl = None


class TestKawtharFileWizard(TransactionCase):
    """Tests for ksw.kawthar.file.wizard."""

    # ================================================================
    # setUp — reusable fixtures
    # ================================================================

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        cls.company = cls.env.company

        # ── Bank ──
        cls.bank_rajhi = cls.env['res.bank'].create({
            'name': 'Al Rajhi Bank',
            'bic': 'RJHISARI',
        })

        # ── Company bank account ──
        cls.company_bank = cls.env['res.partner.bank'].create({
            'acc_number': 'SA71800005786080100T0001',
            'partner_id': cls.company.partner_id.id,
            'bank_id': cls.bank_rajhi.id,
            'x_wps_cic_number': '1389678',
            'x_wps_debit_account': 'SA7180000578608010033217',
            'x_wps_mol_id': '004-110457',
        })

        # ── Work schedule ──
        cls.calendar = cls.env['resource.calendar'].create({
            'name': 'Kawthar Test Calendar',
            'tz': 'Asia/Riyadh',
        })

        # ── Salary rule category refs ──
        cls.cat_basic = cls.env.ref('om_hr_payroll.BASIC')
        cls.cat_hra = cls.env.ref('om_hr_payroll.HRA')
        cls.cat_gross = cls.env.ref('om_hr_payroll.GROSS')
        cls.cat_ded = cls.env.ref('om_hr_payroll.DED')
        cls.cat_net = cls.env.ref('om_hr_payroll.NET')

        # ── Salary rules ──
        cls.rule_basic = cls._get_or_create_rule(cls, 'BASIC', cls.cat_basic)
        cls.rule_hra = cls._get_or_create_rule(cls, 'HRA', cls.cat_hra)
        cls.rule_gross = cls._get_or_create_rule(cls, 'GROSS', cls.cat_gross)
        cls.rule_ded = cls._get_or_create_rule(cls, 'ATTDED', cls.cat_ded)
        cls.rule_net = cls._get_or_create_rule(cls, 'NET', cls.cat_net)

        # ── Employees WITH bank accounts (Kawthar card numbers) ──
        cls.emp_1 = cls._create_employee(
            cls, 'HUSSAIN ABD AL REDHA AL SHAKHS',
            '001', '1020291827',
            acc_number='5689110000130257800',
        )
        cls.emp_2 = cls._create_employee(
            cls, 'MARWA JAFAR H ALGHAFLI',
            '002', '1110744735',
            acc_number='5689110000105257800',
        )

        # ── Employee WITHOUT bank account ──
        cls.emp_no_bank = cls._create_employee(
            cls, 'NO BANK EMPLOYEE',
            '003', '9999999999',
            acc_number=False,
        )

        # ── Assign company bank ──
        for emp in (cls.emp_1, cls.emp_2, cls.emp_no_bank):
            emp.sudo().write({
                'x_salary_bank_account_id': cls.company_bank.id,
            })

        # ── Payslip batch ──
        cls.batch = cls.env['hr.payslip.run'].create({
            'name': 'March 2026',
            'date_start': date(2026, 3, 1),
            'date_end': date(2026, 3, 31),
        })

        # ── Payslips ──
        cls.slip_1 = cls._create_payslip(
            cls, cls.emp_1, cls.batch,
            basic=7000, hra=0, gross=7000, deductions=0, net=7000,
        )
        cls.slip_2 = cls._create_payslip(
            cls, cls.emp_2, cls.batch,
            basic=4000, hra=1000, gross=5488, deductions=488, net=5000,
        )
        cls.slip_no_bank = cls._create_payslip(
            cls, cls.emp_no_bank, cls.batch,
            basic=3000, hra=0, gross=3000, deductions=0, net=3000,
        )

    # ── Factory helpers ──

    def _get_or_create_rule(self, code, category):
        rule = self.env['hr.salary.rule'].search(
            [('code', '=', code)], limit=1,
        )
        if not rule:
            rule = self.env['hr.salary.rule'].create({
                'name': code,
                'code': code,
                'category_id': category.id,
                'sequence': 100,
                'condition_select': 'none',
                'amount_select': 'fix',
                'amount_fix': 0,
            })
        return rule

    def _create_employee(self, name, barcode, id_number, acc_number=False):
        emp = self.env['hr.employee'].create({
            'name': name,
            'barcode': barcode,
            'identification_id': id_number,
            'resource_calendar_id': self.calendar.id,
        })
        if acc_number:
            bank_acc = self.env['res.partner.bank'].create({
                'acc_number': acc_number,
                'partner_id': emp.work_contact_id.id,
                'bank_id': self.bank_rajhi.id,
            })
            emp.sudo().write({'bank_account_ids': [(4, bank_acc.id)]})
        return emp

    def _create_payslip(self, employee, batch, basic, hra, gross,
                        deductions, net):
        slip = self.env['hr.payslip'].create({
            'name': 'Slip %s' % employee.name,
            'employee_id': employee.id,
            'date_from': batch.date_start,
            'date_to': batch.date_end,
            'payslip_run_id': batch.id,
            'version_id': employee.current_version_id.id,
        })
        Line = self.env['hr.payslip.line']
        version_id = employee.current_version_id.id

        def _mk(rule, amount):
            Line.create({
                'slip_id': slip.id,
                'salary_rule_id': rule.id,
                'code': rule.code,
                'name': rule.name,
                'category_id': rule.category_id.id,
                'employee_id': employee.id,
                'version_id': version_id,
                'amount': amount,
                'quantity': 1,
                'rate': 100,
                'sequence': rule.sequence,
            })

        _mk(self.rule_basic, basic)
        _mk(self.rule_hra, hra)
        _mk(self.rule_gross, gross)
        if deductions:
            _mk(self.rule_ded, -abs(deductions))
        _mk(self.rule_net, net)
        return slip

    def _make_wizard(self, batch=None, operation_code='2'):
        """Create a wizard instance for the given batch."""
        batch = batch or self.batch
        return self.env['ksw.kawthar.file.wizard'].create({
            'payslip_run_id': batch.id,
            'operation_code': operation_code,
        })

    def _generate_and_read(self, wiz):
        """Run action_generate and return the openpyxl workbook."""
        result = wiz.action_generate()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        data = base64.b64decode(att.datas)
        return openpyxl.load_workbook(io.BytesIO(data))

    # ================================================================
    # Tests — Wizard basics
    # ================================================================

    def test_wizard_default_operation(self):
        """Default operation code is '2' (Load Funds)."""
        wiz = self.env['ksw.kawthar.file.wizard'].with_context(
            default_payslip_run_id=self.batch.id,
        ).create({})
        self.assertEqual(wiz.operation_code, '2')

    def test_wizard_requires_batch(self):
        """payslip_run_id is mandatory."""
        from psycopg2 import IntegrityError
        with self.assertRaises(IntegrityError), self.cr.savepoint():
            self.env['ksw.kawthar.file.wizard'].create({
                'operation_code': '2',
            })

    # ================================================================
    # Tests — action_open_kawthar_wizard
    # ================================================================

    def test_action_open_kawthar_wizard(self):
        """Batch button returns act_window for the Kawthar wizard."""
        action = self.batch.action_open_kawthar_wizard()
        self.assertEqual(action['type'], 'ir.actions.act_window')
        self.assertEqual(action['res_model'], 'ksw.kawthar.file.wizard')
        self.assertEqual(action['target'], 'new')
        self.assertEqual(
            action['context']['default_payslip_run_id'], self.batch.id,
        )

    # ================================================================
    # Tests — Error handling
    # ================================================================

    def test_error_empty_batch(self):
        """Wizard raises if batch has no payslips."""
        empty_batch = self.env['hr.payslip.run'].create({
            'name': 'Empty', 'date_start': date(2026, 3, 1),
            'date_end': date(2026, 3, 31),
        })
        wiz = self._make_wizard(batch=empty_batch)
        with self.assertRaises(UserError):
            wiz.action_generate()

    def test_error_no_bank_employees(self):
        """Wizard raises when no employees have a bank account."""
        batch = self.env['hr.payslip.run'].create({
            'name': 'No Bank',
            'date_start': date(2026, 3, 1),
            'date_end': date(2026, 3, 31),
        })
        self._create_payslip(
            self.emp_no_bank, batch,
            basic=3000, hra=0, gross=3000, deductions=0, net=3000,
        )
        wiz = self._make_wizard(batch=batch)
        with self.assertRaises(UserError):
            wiz.action_generate()

    # ================================================================
    # Tests — Sheet structure
    # ================================================================

    def test_sheet_name(self):
        """Generated sheet is named 'PREFORMAT PAYMENTS'."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        self.assertIn('PREFORMAT PAYMENTS', wb.sheetnames)

    def test_title_row(self):
        """Row 1 starts with 'Payroll Cards File Upload'."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertEqual(ws.cell(1, 1).value, 'Payroll Cards File Upload')

    def test_notes_row(self):
        """Row 2 contains the notes text."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertIn('Do Not change the order', str(ws.cell(2, 1).value))

    def test_english_headers(self):
        """Row 3 has the correct English column headers."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertEqual(ws.cell(3, 1).value, 'Employee ID (12N)')
        self.assertEqual(ws.cell(3, 2).value, 'Payroll Card No.  (19N)')
        self.assertEqual(ws.cell(3, 3).value, 'Employee Name (50A)')
        self.assertEqual(ws.cell(3, 4).value, 'National ID Number (10N)')
        self.assertEqual(ws.cell(3, 5).value, 'Amount (15N)')
        self.assertEqual(ws.cell(3, 6).value, 'Operating Code (1X)')

    def test_arabic_headers(self):
        """Row 4 has Arabic column headers."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertEqual(ws.cell(4, 1).value, 'الرقم الوظيفي')
        self.assertEqual(ws.cell(4, 2).value, 'رقم بطاقة اتقان')

    def test_operation_type_columns(self):
        """Row 1 has operation-type labels starting at column P (16)."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertEqual(ws.cell(1, 16).value, 'Create New Card')
        self.assertEqual(ws.cell(1, 17).value, 'Load Funds')

    # ================================================================
    # Tests — Data rows
    # ================================================================

    def test_data_row_count(self):
        """Only employees with bank accounts and positive NET are included."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # 4 header rows + 2 data rows (emp_no_bank excluded)
        data_rows = 0
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 1).value is not None:
                data_rows += 1
        self.assertEqual(data_rows, 2)

    def test_sequential_employee_id(self):
        """Employee ID column uses sequential numbering (1, 2, 3…)."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertEqual(ws.cell(5, 1).value, 1)
        self.assertEqual(ws.cell(6, 1).value, 2)

    def test_card_number_from_bank_account(self):
        """Payroll Card No. column uses primary_bank_account_id.acc_number."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # Sorted alphabetically: HUSSAIN < MARWA
        self.assertEqual(
            ws.cell(5, 2).value, '5689110000130257800',
        )
        self.assertEqual(
            ws.cell(6, 2).value, '5689110000105257800',
        )

    def test_employee_name_uppercase(self):
        """Employee name is uppercased."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertEqual(
            ws.cell(5, 3).value, 'HUSSAIN ABD AL REDHA AL SHAKHS',
        )

    def test_national_id(self):
        """National ID column has the employee's identification_id."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertEqual(ws.cell(5, 4).value, '1020291827')

    def test_amount_whole_sar(self):
        """Amount column is in whole SAR (not halalas)."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # emp_1 net=7000, emp_2 net=5000
        self.assertEqual(ws.cell(5, 5).value, 7000)
        self.assertEqual(ws.cell(6, 5).value, 5000)

    def test_operation_code_label(self):
        """Operating Code column shows the Arabic label for selected code."""
        wiz = self._make_wizard(operation_code='2')
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertIn('تحميل رصيد', ws.cell(5, 6).value)

    def test_operation_code_create_card(self):
        """Operation code '1' maps to 'أنشاء بطاقة جديدة'."""
        wiz = self._make_wizard(operation_code='1')
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        self.assertIn('أنشاء بطاقة جديدة', ws.cell(5, 6).value)

    # ================================================================
    # Tests — Salary breakdown
    # ================================================================

    def test_basic_salary(self):
        """Basic Salary column (H) has the correct value."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # emp_1 basic=7000
        self.assertEqual(ws.cell(5, 8).value, 7000)
        # emp_2 basic=4000
        self.assertEqual(ws.cell(6, 8).value, 4000)

    def test_housing_allowance(self):
        """Housing Allowance column (I) has the correct value."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # emp_1 hra=0, emp_2 hra=1000
        self.assertEqual(ws.cell(5, 9).value, 0)
        self.assertEqual(ws.cell(6, 9).value, 1000)

    def test_other_allowance(self):
        """Other Allowance = GROSS - BASIC - HRA."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # emp_1: 7000 - 7000 - 0 = 0
        self.assertEqual(ws.cell(5, 10).value, 0)
        # emp_2: 5488 - 4000 - 1000 = 488
        self.assertEqual(ws.cell(6, 10).value, 488)

    def test_deductions(self):
        """Deductions column (K) has the absolute deduction value."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # emp_1 deductions=0, emp_2 deductions=488
        self.assertEqual(ws.cell(5, 11).value, 0)
        self.assertEqual(ws.cell(6, 11).value, 488)

    def test_salary_formula(self):
        """basic + hra + other - deductions == net."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        for r in range(5, 7):
            net = ws.cell(r, 5).value
            basic = ws.cell(r, 8).value
            hra = ws.cell(r, 9).value
            other = ws.cell(r, 10).value
            ded = ws.cell(r, 11).value
            self.assertEqual(
                basic + hra + other - ded, net,
                'Salary formula mismatch on row %d' % r,
            )

    # ================================================================
    # Tests — Zero-net slips excluded
    # ================================================================

    def test_zero_net_slips_excluded(self):
        """Payslips with NET=0 are excluded."""
        zero_emp = self._create_employee(
            'ZERO SALARY', '999', '0000000000',
            acc_number='5689110000999957800',
        )
        zero_emp.sudo().write({
            'x_salary_bank_account_id': self.company_bank.id,
        })
        self._create_payslip(
            zero_emp, self.batch,
            basic=0, hra=0, gross=0, deductions=0, net=0,
        )
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        data_rows = 0
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 1).value is not None:
                data_rows += 1
        self.assertEqual(data_rows, 2)

    # ================================================================
    # Tests — File output
    # ================================================================

    def test_action_generate_returns_download_url(self):
        """action_generate returns an ir.actions.act_url for download."""
        wiz = self._make_wizard()
        result = wiz.action_generate()
        self.assertEqual(result['type'], 'ir.actions.act_url')
        self.assertIn('/web/content/', result['url'])
        self.assertIn('download=true', result['url'])

    def test_action_generate_creates_attachment(self):
        """An ir.attachment is created for the batch."""
        wiz = self._make_wizard()
        before = self.env['ir.attachment'].search_count([
            ('res_model', '=', 'hr.payslip.run'),
            ('res_id', '=', self.batch.id),
        ])
        wiz.action_generate()
        after = self.env['ir.attachment'].search_count([
            ('res_model', '=', 'hr.payslip.run'),
            ('res_id', '=', self.batch.id),
        ])
        self.assertEqual(after - before, 1)

    def test_attachment_filename(self):
        """Attachment filename includes 'Kawthar' and batch name."""
        wiz = self._make_wizard()
        wiz.action_generate()
        att = self.env['ir.attachment'].search([
            ('res_model', '=', 'hr.payslip.run'),
            ('res_id', '=', self.batch.id),
            ('name', 'like', 'Kawthar_%'),
        ], order='id desc', limit=1)
        self.assertIn('Kawthar', att.name)
        self.assertIn('March_2026', att.name)
        self.assertTrue(att.name.endswith('.xls'))

    # ================================================================
    # Tests — Sorting
    # ================================================================

    def test_data_rows_sorted_by_name(self):
        """Data rows are sorted alphabetically by employee name."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        name_1 = ws.cell(5, 3).value
        name_2 = ws.cell(6, 3).value
        self.assertLess(name_1, name_2)

    # ================================================================
    # Tests — No-bank employees are excluded (not errored)
    # ================================================================

    def test_no_bank_employee_excluded(self):
        """Employees without a bank account are excluded from file."""
        wiz = self._make_wizard()
        wb = self._generate_and_read(wiz)
        ws = wb['PREFORMAT PAYMENTS']
        # Check no row contains 'NO BANK EMPLOYEE'
        for r in range(5, ws.max_row + 1):
            name = ws.cell(r, 3).value or ''
            self.assertNotIn('NO BANK', name)

    # ================================================================
    # Tests — Kawthar TXT format (_build_kawthar_text)
    # ================================================================

    def test_txt_line_length(self):
        """Each Kawthar TXT line is exactly 194 characters."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank,
            self.slip_1 | self.slip_2,
            date(2026, 3, 30),
        )
        lines = [l for l in content.split('\n') if l]
        for line in lines:
            self.assertEqual(len(line), 194)

    def test_txt_no_header(self):
        """Kawthar TXT has no header row — each line is data."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        lines = [l for l in content.split('\n') if l]
        self.assertEqual(len(lines), 1)

    def test_txt_employee_barcode_field(self):
        """First 12 chars are zero-padded employee barcode."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        self.assertEqual(line[:12], '000000000001')

    def test_txt_cic_field(self):
        """Chars 13-22 are CIC number zero-padded to 10 digits."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        self.assertEqual(line[12:22], '0001389678')

    def test_txt_card_number_field(self):
        """Chars 23-36 are card number (acc_number) left-justified 14 chars."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        card = line[22:36]
        self.assertEqual(len(card), 14)
        self.assertTrue(card.startswith('56891100001302'))

    def test_txt_name_field(self):
        """Chars 37-86 are the employee name left-justified 50 chars."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        name = line[36:86]
        self.assertEqual(len(name), 50)
        self.assertTrue(name.startswith('HUSSAIN ABD AL REDHA AL SHAKHS'))

    def test_txt_national_id_field(self):
        """Chars 87-96 are national ID zero-padded 10 digits."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        self.assertEqual(line[86:96], '1020291827')

    def test_txt_net_in_halalas(self):
        """Chars 97-111 are net salary in halalas (15 digits)."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        # 7000 SAR = 700000 halalas
        self.assertEqual(line[96:111], '000000000700000')

    def test_txt_value_date(self):
        """Chars 112-119 are the value date YYYYMMDD."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        self.assertEqual(line[111:119], '20260330')

    def test_txt_operation_code(self):
        """Char 120 is the operation code."""
        wiz = self._make_wizard(operation_code='2')
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        self.assertEqual(line[119], '2')

    def test_txt_zeros_filler(self):
        """Chars 121-126 are 6 zeros."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        self.assertEqual(line[120:126], '000000')

    def test_txt_spaces_filler(self):
        """Chars 127-146 are 20 spaces."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_1, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        self.assertEqual(line[126:146], ' ' * 20)

    def test_txt_salary_breakdown(self):
        """Basic/HRA/Other/Deductions in halalas at end of line."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_2, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        # emp_2: basic=4000, hra=1000, gross=5488, ded=488, net=5000
        basic = int(line[146:158])  # 12 chars
        hra = int(line[158:170])    # 12 chars
        other = int(line[170:182])  # 12 chars
        ded = int(line[182:194])    # 12 chars
        # basic = 4000 SAR = 400000 halalas
        self.assertEqual(basic, 400000)
        # hra = 1000 SAR = 100000 halalas
        self.assertEqual(hra, 100000)
        # other = 5488 - 4000 - 1000 = 488 SAR = 48800 halalas
        self.assertEqual(other, 48800)
        # ded = 488 SAR = 48800 halalas
        self.assertEqual(ded, 48800)

    def test_txt_salary_formula(self):
        """basic + hra + other - ded == net in halalas."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank, self.slip_2, date(2026, 3, 30),
        )
        line = content.split('\n')[0]
        net_h = int(line[96:111])
        basic_h = int(line[146:158])
        hra_h = int(line[158:170])
        other_h = int(line[170:182])
        ded_h = int(line[182:194])
        self.assertEqual(basic_h + hra_h + other_h - ded_h, net_h)

    def test_txt_excludes_zero_net(self):
        """Zero-net slips are excluded from TXT output."""
        wiz = self._make_wizard()
        zero_emp = self._create_employee(
            'ZERO TXT', '888', '8888888888',
            acc_number='5689110000888857800',
        )
        zero_emp.sudo().write({
            'x_salary_bank_account_id': self.company_bank.id,
        })
        zero_slip = self._create_payslip(
            zero_emp, self.batch,
            basic=0, hra=0, gross=0, deductions=0, net=0,
        )
        content = wiz._build_kawthar_text(
            self.company_bank,
            self.slip_1 | self.slip_2 | zero_slip,
            date(2026, 3, 30),
        )
        lines = [l for l in content.split('\n') if l]
        # Only 2 lines (zero excluded)
        self.assertEqual(len(lines), 2)

    def test_txt_sorted_by_name(self):
        """TXT lines are sorted alphabetically by employee name."""
        wiz = self._make_wizard()
        content = wiz._build_kawthar_text(
            self.company_bank,
            self.slip_1 | self.slip_2,
            date(2026, 3, 30),
        )
        lines = [l for l in content.split('\n') if l]
        name_1 = lines[0][36:86].strip()
        name_2 = lines[1][36:86].strip()
        # HUSSAIN < MARWA alphabetically
        self.assertLess(name_1, name_2)

