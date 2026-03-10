# Copyright 2017 Creu Blanca
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html).

import io
import logging

from odoo.tests import common

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    _logger.debug("Can not import openpyxl`.")


class TestReport(common.TransactionCase):
    def setUp(self):
        super().setUp()
        self.report_object = self.env["ir.actions.report"]
        vals = {
            "name": "Print to XLSX",
            "model": "res.partner",
            "report_type": "xlsx",
            "report_name": "report_xlsx.partner_xlsx",
            "report_file": "res_partner",
        }
        self.xlsx_report = self.env["report.report_xlsx.abstract"].with_context(
            active_model="res.partner"
        )
        self.report_name = "report_xlsx.partner_xlsx"
        self.report = self.report_object.create(vals)
        self.docs = self.env["res.company"].search([], limit=1).partner_id

    def test_report(self):
        report = self.report
        self.assertEqual(report.report_type, "xlsx")
        rep = self.report_object._render(self.report, self.docs.ids, {})
        file = io.BytesIO(rep[0])
        wb = load_workbook(file)
        sheet = wb.active
        self.assertEqual(sheet.cell(1, 1).value, self.docs.name)

    def test_save_attachment(self):
        self.report.attachment = 'object.name + ".xlsx"'
        self.report_object._render(self.report, self.docs.ids, {})
        attachment = self.env["ir.attachment"].search(
            [("res_id", "=", self.docs.id), ("res_model", "=", self.docs._name)]
        )
        self.assertEqual(len(attachment), 1)
        self.assertEqual(attachment.name, f"{self.docs.name}.xlsx")

    def test_id_retrieval(self):
        # Typical call from WebUI with wizard
        objs = self.xlsx_report._get_objs_for_report(
            False, {"context": {"active_ids": self.docs.ids}}
        )
        self.assertEqual(objs, self.docs)

        # Typical call from within code not to report_action
        objs = self.xlsx_report.with_context(
            active_ids=self.docs.ids
        )._get_objs_for_report(False, False)
        self.assertEqual(objs, self.docs)

        # Typical call from WebUI
        objs = self.xlsx_report._get_objs_for_report(
            self.docs.ids, {"data": [self.report_name, self.report.report_type]}
        )
        self.assertEqual(objs, self.docs)

        # Typical call from render
        objs = self.xlsx_report._get_objs_for_report(self.docs.ids, {})
        self.assertEqual(objs, self.docs)

    def test_currency_format(self):
        usd = self.env.ref("base.USD")
        self.assertEqual(
            self.xlsx_report._report_xlsx_currency_format(usd), "$#,##0.00"
        )
        eur = self.env.ref("base.EUR")
        self.assertEqual(
            self.xlsx_report._report_xlsx_currency_format(eur), "#,##0.00 €"
        )
