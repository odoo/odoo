# -*- coding: utf-8 -*-
import requests
from odoo.exceptions import UserError
from .msal_helper import acquire_token

# Correct openpyxl imports
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter,
)

class GraphClient:
    BASE_URL = "https://graph.microsoft.com/v1.0"

    def __init__(self, env, file_id, sheet_name="Sheet1"):
        self.env = env
        params = env['ir.config_parameter'].sudo()
        self.token = params.get_param('excel_export.access_token')
        self.drive_id = params.get_param('excel_export.drive_id')
        self.file_id = file_id
        self.sheet = sheet_name

        if not (self.token and self.drive_id and self.file_id):
            raise UserError("Please configure connector and click “Generate Token & Drive ID” first.")

        self.headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json",
        }

    def _request(self, method, url, **kwargs):
        resp = requests.request(method, url, headers=self.headers, **kwargs)
        if resp.status_code == 401:
            # Token expired → fetch new one & retry once
            new_token = acquire_token(self.env)
            self.env['ir.config_parameter'].sudo().set_param('excel_export.access_token', new_token)
            self.headers["Authorization"] = f"Bearer {new_token}"
            resp = requests.request(method, url, headers=self.headers, **kwargs)
        try:
            resp.raise_for_status()
        except requests.HTTPError:
            # Show real error from Microsoft Graph for debugging
            raise UserError(f"Graph API error {resp.status_code}: {resp.text}")
        return resp

    def clear_sheet(self):
        url = (
            f"{self.BASE_URL}/drives/{self.drive_id}"
            f"/items/{self.file_id}"
            f"/workbook/worksheets('{self.sheet}')"
            f"/usedRange/clear"
        )
        self._request("POST", url, json={"applyTo": "All"})

    def get_worksheet_names(self):
        """Helper: List all worksheet names for debugging."""
        url = (
            f"{self.BASE_URL}/drives/{self.drive_id}"
            f"/items/{self.file_id}/workbook/worksheets"
        )
        resp = self._request("GET", url)
        data = resp.json()
        return [ws['name'] for ws in data.get('value', [])]

    def _get_range_address(self, start_cell, num_rows, num_cols):
        """
        Calculate the Excel range address (e.g., A1:C3) for the given data shape.
        """
        col, row = coordinate_from_string(start_cell)
        start_col_idx = column_index_from_string(col)
        end_col_idx = start_col_idx + num_cols - 1
        end_col_letter = get_column_letter(end_col_idx)
        end_row_num = row + num_rows - 1
        return f"{col}{row}:{end_col_letter}{end_row_num}"

    def write_values(self, values, start_cell="A1"):
        # Ensure values is a 2D list
        if not (isinstance(values, list) and all(isinstance(row, list) for row in values)):
            raise UserError("Values must be a list of lists (2D array) for Excel API.")

        num_rows = len(values)
        num_cols = len(values[0]) if values else 0
        if num_rows == 0 or num_cols == 0:
            raise UserError("Values must be a non-empty 2D array.")

        range_address = self._get_range_address(start_cell, num_rows, num_cols)
        url = (
            f"{self.BASE_URL}/drives/{self.drive_id}"
            f"/items/{self.file_id}"
            f"/workbook/worksheets('{self.sheet}')"
            f"/range(address='{range_address}')"
        )
        self._request("PATCH", url, json={"values": values})
