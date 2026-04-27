# Part of Odoo. See LICENSE file for full copyright and licensing details.

import io
import unittest
from datetime import datetime
from freezegun import freeze_time
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


from odoo import Command
from odoo.addons.account_reports.tests.account_sales_report_common import AccountSalesReportCommon
from odoo.tests import tagged


@tagged('post_install_l10n', 'post_install', '-at_install')
class L10nThaiTaxReportTest(AccountSalesReportCommon):

    @classmethod
    @AccountSalesReportCommon.setup_country('th')
    def setUpClass(cls):
        super().setUpClass()
        cls.partner_b.write({
            'country_id': cls.env.ref('base.th').id,
            "vat": "1234545678781",
            "company_registry": "12345678"
        })

    def compare_xlsx_data(self, report_data, expected_data):
        if load_workbook is None:
            raise unittest.SkipTest("openpyxl not available")
        report_file = io.BytesIO(report_data)
        xlsx = load_workbook(filename=report_file, data_only=True)
        sheet = xlsx.worksheets[0]
        sheet_values = list(sheet.values)

        result = []

        for row in sheet_values[6:]:
            result.append([v if v is not None else '' for v in list(row)])

        self.assertEqual(result, expected_data)

    @freeze_time('2023-06-30')
    def test_pnd53_report(self):
        tax_1 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_co_3')
        tax_2 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_co_2')

        move = self.env['account.move'].create({
            'move_type': 'in_invoice',
            'journal_id': self.company_data['default_journal_purchase'].id,
            'partner_id': self.partner_b.id,
            'invoice_date': '2023-05-20',
            'date': '2023-05-20',
            'company_id': self.company_data['company'].id,
            'invoice_line_ids': [
                (0, 0, {
                    'product_id': self.product_a.id,
                    'quantity': 1,
                    'name': 'product test 1',
                    'price_unit': 1000,
                    'tax_ids': tax_1.ids
                }),
                (0, 0, {
                    'product_id': self.product_b.id,
                    'quantity': 1,
                    'name': 'product test 2',
                    'price_unit': 1000,
                    'tax_ids': tax_2.ids
                })
            ]
        })
        move.action_post()
        self.env.flush_all()

        report = self.env.ref('l10n_th.tax_report_pnd53')
        options = report.get_options({})

        report_data = self.env['l10n_th.pnd53.report.handler'].l10n_th_print_pnd_tax_report_pnd53(options)['file_content']
        expected = ("No.,Tax ID,Title,Contact Name,Street,Street2,City,State,Zip,Branch Number,Invoice/Bill Date,Tax Rate,Total Amount,WHT Amount,WHT Condition,Tax Type\n"
                    f"1,{self.partner_b.vat},บริษัท,Partner B,,,,,,12345678,20/05/2023,3.00,1000.00,30.00,1,Service\n"
                    f"2,{self.partner_b.vat},บริษัท,Partner B,,,,,,12345678,20/05/2023,2.00,1000.00,20.00,1,Advertising\n").encode()

        self.assertEqual(report_data, expected)

    @freeze_time('2023-06-30')
    def test_pnd3_report(self):
        tax_1 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_pers_1')
        tax_2 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_pers_2')

        move = self.env['account.move'].create({
            'move_type': 'in_invoice',
            'journal_id': self.company_data['default_journal_purchase'].id,
            'partner_id': self.partner_b.id,
            'invoice_date': '2023-05-20',
            'date': '2023-05-20',
            'company_id': self.company_data['company'].id,
            'invoice_line_ids': [
                (0, 0, {
                    'product_id': self.product_a.id,
                    'quantity': 1,
                    'name': 'product test 1',
                    'price_unit': 1000,
                    'tax_ids': tax_1.ids
                }),
                (0, 0, {
                    'product_id': self.product_b.id,
                    'quantity': 1,
                    'name': 'product test 2',
                    'price_unit': 1000,
                    'tax_ids': tax_2.ids
                })
            ]
        })
        move.action_post()
        self.env.flush_all()

        report = self.env.ref('l10n_th.tax_report_pnd3')
        options = report.get_options({})

        report_data = self.env['l10n_th.pnd3.report.handler'].l10n_th_print_pnd_tax_report_pnd3(options)['file_content']
        expected = ("No.,Tax ID,Title,Contact Name,Street,Street2,City,State,Zip,Branch Number,Invoice/Bill Date,Tax Rate,Total Amount,WHT Amount,WHT Condition,Tax Type\n"
                    f"1,{self.partner_b.vat},,Partner B,,,,,,12345678,20/05/2023,1.00,1000.00,10.00,1,Transportation\n"
                    f"2,{self.partner_b.vat},,Partner B,,,,,,12345678,20/05/2023,2.00,1000.00,20.00,1,Advertising\n").encode()

        self.assertEqual(report_data, expected)

    @freeze_time('2023-06-30')
    def test_vat_tax_report_branch_name(self):
        self.partner_b.is_company = True
        tax_1 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_co_3')
        tax_2 = self.env.ref(f'account.{self.company_data["company"].id}_tax_output_vat')
        self.init_invoice("out_invoice", self.partner_b, "2023-05-20", amounts=[1000, 1000], taxes=[tax_1, tax_2], post=True)

        report = self.env.ref('l10n_th.tax_report')
        options = report.get_options({})

        report_data = self.env['l10n_th.tax.report.handler'].l10n_th_print_sale_tax_report(options)['file_content']
        expected = [
            ['No.', 'Tax Invoice No.', 'Reference', 'Invoice Date', 'Contact Name', 'Tax ID', 'Company Information', 'Total Amount', 'Total Excluding VAT Amount', 'Vat Amount'],
            [1, 'INV/2023/00001', '', datetime(2023, 5, 20), 'Partner B', self.partner_b.vat, 'Branch 12345678', 2080.0, 2000.0, 140.0],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', 'Total', 2080.0, 2000.0, 140]
        ]
        self.compare_xlsx_data(report_data, expected)

    @freeze_time('2023-06-30')
    def test_vat_sales_tax_report(self):
        tax_1 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_co_3')
        tax_2 = self.env.ref(f'account.{self.company_data["company"].id}_tax_output_vat')
        self.init_invoice("out_invoice", self.partner_b, "2023-05-20", amounts=[1000, 1000], taxes=[tax_1, tax_2], post=True)

        report = self.env.ref('l10n_th.tax_report')
        options = report.get_options({})

        report_data = self.env['l10n_th.tax.report.handler'].l10n_th_print_sale_tax_report(options)['file_content']
        expected = [
            ['No.', 'Tax Invoice No.', 'Reference', 'Invoice Date', 'Contact Name', 'Tax ID', 'Company Information', 'Total Amount', 'Total Excluding VAT Amount', 'Vat Amount'],
            [1, 'INV/2023/00001', '', datetime(2023, 5, 20), 'Partner B', self.partner_b.vat, '', 2080.0, 2000.0, 140.0],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', 'Total', 2080.0, 2000.0, 140]
        ]
        self.compare_xlsx_data(report_data, expected)

    @freeze_time('2023-06-30')
    def test_credit_note_sales_tax_report(self):
        tax_1 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_co_3')
        tax_2 = self.env.ref(f'account.{self.company_data["company"].id}_tax_output_vat')
        self.init_invoice("out_refund", self.partner_b, "2023-05-21", amounts=[500, 500], taxes=[tax_1, tax_2], post=True)

        report = self.env.ref('l10n_th.tax_report')
        options = report.get_options({})

        report_data = self.env['l10n_th.tax.report.handler'].l10n_th_print_sale_tax_report(options)['file_content']
        expected = [
            ['No.', 'Tax Invoice No.', 'Reference', 'Invoice Date', 'Contact Name', 'Tax ID', 'Company Information', 'Total Amount', 'Total Excluding VAT Amount', 'Vat Amount'],
            [1, 'RINV/2023/00001', '', datetime(2023, 5, 21), 'Partner B', self.partner_b.vat, '', -1040.0, -1000.0, -70.0],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', 'Total', -1040, -1000.0, -70]
        ]
        self.compare_xlsx_data(report_data, expected)

    @freeze_time('2023-06-30')
    def test_vat_purchase_tax_report_full_refund(self):
        tax_1 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_co_3')
        tax_2 = self.env.ref(f'account.{self.company_data["company"].id}_tax_input_vat')
        self.init_invoice("in_invoice", self.partner_b, "2023-05-20", amounts=[1000, 1000], taxes=[tax_1, tax_2], post=True)

        # Reversed move should not be included in the report
        move_to_reverse = self.init_invoice("in_invoice", self.partner_b, "2023-05-20", amounts=[1000, 1000], taxes=[tax_1, tax_2], post=True)
        move_to_reverse._reverse_moves([{"invoice_date": move_to_reverse.date}], cancel=True)

        report = self.env.ref('l10n_th.tax_report')
        options = report.get_options({})

        report_data = self.env['l10n_th.tax.report.handler'].l10n_th_print_purchase_tax_report(options)['file_content']
        expected = [
            ['No.', 'Tax Invoice No.', 'Reference', 'Invoice Date', 'Contact Name', 'Tax ID', 'Company Information', 'Total Amount', 'Total Excluding VAT Amount', 'Vat Amount'],
            [1, 'BILL/2023/05/0001', '', datetime(2023, 5, 20), 'Partner B', self.partner_b.vat, '', 2080.0, 2000.0, 140.0],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', 'Total', 2080.0, 2000.0, 140.0]
        ]
        self.compare_xlsx_data(report_data, expected)

    @freeze_time('2023-06-30')
    def test_vat_purchase_tax_report_partial_refund(self):
        tax_1 = self.env.ref(f'account.{self.company_data["company"].id}_tax_wht_co_3')
        tax_2 = self.env.ref(f'account.{self.company_data["company"].id}_tax_input_vat')
        move_to_partial_refund = self.init_invoice("in_invoice", self.partner_b, "2023-05-20", amounts=[1000, 1000], taxes=[tax_1, tax_2], post=True)
        reverse_move = move_to_partial_refund._reverse_moves([{"invoice_date": move_to_partial_refund.date, "date": move_to_partial_refund.date}])
        reverse_move.write({'invoice_line_ids': [
            Command.update(line.id, {
                'price_unit': 500,
            }) for line in reverse_move.invoice_line_ids
        ]})
        reverse_move.action_post()
        report = self.env.ref('l10n_th.tax_report')
        options = report.get_options({})

        report_data = self.env['l10n_th.tax.report.handler'].l10n_th_print_purchase_tax_report(options)['file_content']
        expected = [
            ['No.', 'Tax Invoice No.', 'Reference', 'Invoice Date', 'Contact Name', 'Tax ID', 'Company Information', 'Total Amount', 'Total Excluding VAT Amount', 'Vat Amount'],
            [1, 'RBILL/2023/05/0001', '', datetime(2023, 5, 20), 'Partner B', self.partner_b.vat, '', -1040.0, -1000.0, -70.0],
            [2, 'BILL/2023/05/0001', '', datetime(2023, 5, 20), 'Partner B', self.partner_b.vat, '', 2080.0, 2000.0, 140.0],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', '', '', '', ''],
            ['', '', '', '', '', '', 'Total', 1040.0, 1000.0, 70.0]
        ]
        self.compare_xlsx_data(report_data, expected)
