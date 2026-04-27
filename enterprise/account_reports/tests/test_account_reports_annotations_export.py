import datetime
import io
import unittest

from odoo import Command
from odoo.tests import tagged
from odoo.addons.account_reports.tests.common import TestAccountReportsCommon

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@tagged('post_install', '-at_install')
class TestAccountReportAnnotationsExport(TestAccountReportsCommon):
    @classmethod
    def setUpClass(cls):
        if load_workbook is None:
            raise unittest.SkipTest("openpyxl not available")

        super().setUpClass()

        cls.report = cls.env.ref('account_reports.balance_sheet')
        cls.report.column_ids.sortable = True

        # Get accounts
        bank_default_account = cls.company_data["default_journal_bank"].default_account_id
        tax_sale_default_account = cls.company_data["default_account_tax_sale"]
        # Create move
        move = cls.env['account.move'].create({
            'move_type': 'entry',
            'date': '2024-06-10',
            'journal_id': cls.company_data['default_journal_cash'].id,
            'line_ids': [
                (0, 0, {'debit':  250.0,     'credit':   0.0,    'account_id': bank_default_account.id}),
                (0, 0, {'debit':   0.0,     'credit': 250.0,    'account_id': tax_sale_default_account.id}),
            ],
        })
        move.action_post()
        # Get line_ids
        line_id_ta = cls.report._get_generic_line_id('account.report.line', cls.env.ref('account_reports.account_financial_report_total_assets0').id)
        line_id_ca = cls.report._get_generic_line_id('account.report.line', cls.env.ref('account_reports.account_financial_report_current_assets_view0').id, parent_line_id=line_id_ta)
        line_id_ba = cls.report._get_generic_line_id('account.report.line', cls.env.ref('account_reports.account_financial_report_bank_view0').id, parent_line_id=line_id_ca)
        line_id_bank = cls.report._get_generic_line_id('account.account', bank_default_account.id, markup={'groupby': 'account_id'}, parent_line_id=line_id_ba)
        # Create annotation
        date = datetime.datetime.strptime('2024-06-20', '%Y-%m-%d').date()
        cls.report.write({
            'annotations_ids': [
                Command.create({
                    'line_id': cls.env['account.report.annotation']._remove_tax_grouping_from_line_id(line_id_bank),
                    'text': 'Papa a vu le fifi de lolo',
                    'date': date,
                }),
            ]
        })

    def read_xlsx_data(self, report_data):
        report_file = io.BytesIO(report_data)
        xlsx = load_workbook(filename=report_file, data_only=True)
        sheet = xlsx.worksheets[0]
        return list(sheet.values)

    def test_annotations_export_no_comparison(self):
        options = self._generate_options(self.report, '2024-06-01', '2024-06-30', default_options={'unfold_all': True})
        report_data = self.report.export_to_xlsx(options)
        export_content = self.read_xlsx_data(report_data['file_content'])

        # When there is no comparison, there are two columns for the accounts (number, name), one column for the data and
        # finally one column for the annotations. Hence the index 3 for the annotation column.
        self.assertEqual(export_content[0][3], "Annotations")
        self.assertEqual(export_content[7][3], "1 - Papa a vu le fifi de lolo")

    def test_annotations_export_same_period_last_year_comparison(self):
        default_comparison = {
            'filter': 'same_last_year',
            'number_period': 1,
            'date_from': '2023-06-01',
            'date_to': '2023-06-30',
            'periods': [{
                'string': 'As of 06/30/2023',
                'period_type': 'month', 'mode':
                'single', 'date_from': '2023-06-01',
                'date_to': '2023-06-30'
            }],
            'period_order':
            'descending',
            'string': 'As of 06/30/2023',
            'period_type': 'month',
            'mode': 'single'
        }
        options = self._generate_options(self.report, '2024-06-01', '2024-06-30', default_options={'unfold_all': True, 'comparison': default_comparison})
        report_data = self.report.export_to_xlsx(options)
        export_content = self.read_xlsx_data(report_data['file_content'])

        # When comparing with the same period last year, there are two columns for the accounts (number, name), two columns for the
        # date (one per period), one column for the growth comparison percentage and finally one column for the annotations.
        # Hence the index 5 for the annotation column.
        self.assertEqual(export_content[0][5], "Annotations")
        self.assertEqual(export_content[7][5], "1 - Papa a vu le fifi de lolo")

    def test_annotations_export_two_last_periods_comparison(self):
        default_comparison = {
            'filter': 'previous_period',
            'number_period': 2,
            'date_from': '2024-05-01',
            'date_to': '2024-05-31',
            'periods': [
                {
                    'string': 'As of 05/31/2024',
                    'period_type': 'month',
                    'mode': 'single',
                    'date_from': '2024-05-01',
                    'date_to': '2024-05-31'
                },
                {
                    'string': 'As of 04/30/2024',
                    'period_type': 'month',
                    'mode': 'single',
                    'date_from': '2024-04-01',
                    'date_to': '2024-04-30'
                }
            ],
            'period_order': 'descending',
            'string': 'As of 05/31/2024',
            'period_type': 'month',
            'mode': 'single'
        }
        options = self._generate_options(self.report, '2024-06-01', '2024-06-30', default_options={'unfold_all': True, 'comparison': default_comparison})
        report_data = self.report.export_to_xlsx(options)
        export_content = self.read_xlsx_data(report_data['file_content'])

        # When comparing with the two last periods, there are two columns for the accounts (number, name), three columns for the
        # data (one per period) and finally one column for the annotations. Hence the index 5 for the annotation column.
        self.assertEqual(export_content[0][5], "Annotations")
        self.assertEqual(export_content[7][5], "1 - Papa a vu le fifi de lolo")
