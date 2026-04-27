# Part of Odoo. See LICENSE file for full copyright and licensing details.

import copy
import io
import unittest
from collections import Counter
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from odoo.addons.account.tests.common import AccountTestInvoicingCommon

from odoo import Command, fields
from odoo.exceptions import UserError
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT
from odoo.tools.misc import formatLang, file_open

class TestAccountReportsCommon(AccountTestInvoicingCommon):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.other_currency = cls.setup_other_currency('CAD')
        cls.company_data_2 = cls.setup_other_company()
        cls.company_data_2['company'].currency_id = cls.other_currency
        cls.company_data_2['currency'] = cls.other_currency

    @classmethod
    def _generate_options(cls, report, date_from, date_to, default_options=None):
        ''' Create new options at a certain date.
        :param report:          The report.
        :param date_from:       A datetime object, str representation of a date or False.
        :param date_to:         A datetime object or str representation of a date.
        :return:                The newly created options.
        '''
        if isinstance(date_from, datetime):
            date_from_str = fields.Date.to_string(date_from)
        else:
            date_from_str = date_from

        if isinstance(date_to, datetime):
            date_to_str = fields.Date.to_string(date_to)
        else:
            date_to_str = date_to

        if not default_options:
            default_options = {}

        return report.get_options({
            'selected_variant_id': report.id,
            'date': {
                'date_from': date_from_str,
                'date_to': date_to_str,
                'mode': 'range',
                'filter': 'custom',
            },
            'show_account': True,
            'show_currency': True,
            **default_options,
        })

    def _update_comparison_filter(self, options, report, comparison_type, number_period, date_from=None, date_to=None):
        ''' Modify the existing options to set a new filter_comparison.
        :param options:         The report options.
        :param report:          The report.
        :param comparison_type: One of the following values: ('no_comparison', 'custom', 'previous_period', 'previous_year').
        :param number_period:   The number of period to compare.
        :param date_from:       A datetime object for the 'custom' comparison_type.
        :param date_to:         A datetime object the 'custom' comparison_type.
        :return:                The newly created options.
        '''
        previous_options = {**options, 'comparison': {
            **options['comparison'],
            'date_from': date_from and date_from.strftime(DEFAULT_SERVER_DATE_FORMAT),
            'date_to': date_to and date_to.strftime(DEFAULT_SERVER_DATE_FORMAT),
            'filter': comparison_type,
            'number_period': number_period,
        }}
        return report.get_options(previous_options)

    def _update_multi_selector_filter(self, options, option_key, selected_ids):
        ''' Modify a selector in the options to select .
        :param options:         The report options.
        :param option_key:      The key to the option.
        :param selected_ids:    The ids to be selected.
        :return:                The newly created options.
        '''
        new_options = copy.deepcopy(options)
        for c in new_options[option_key]:
            c['selected'] = c['id'] in selected_ids
        return new_options

    def assertColumnPercentComparisonValues(self, lines, expected_values):
        filtered_lines = self._filter_folded_lines(lines)

        # Check number of lines.
        self.assertEqual(len(filtered_lines), len(expected_values))

        for value, expected_value in zip(filtered_lines, expected_values):
            # Check number of columns.
            key = 'column_percent_comparison_data'
            self.assertEqual(len(value[key]) + 1, len(expected_value))
            # Check name, value and class.
            self.assertEqual((value['name'], value[key]['name'], value[key]['mode']), expected_value)

    def assertHorizontalGroupTotal(self, lines, expected_values):
        filtered_lines = self._filter_folded_lines(lines)

        # Check number of lines.
        self.assertEqual(len(filtered_lines), len(expected_values))
        for line_dict_list, expected_values in zip(filtered_lines, expected_values):
            column_values = [column['no_format'] for column in line_dict_list['columns']]
            # Compare the Total column, Total column is there only under certain condition
            if line_dict_list.get('horizontal_group_total_data'):
                self.assertEqual(len(line_dict_list['columns']) + 1, len(expected_values[1:]))
                # Compare the numbers column except the total
                self.assertEqual(column_values, list(expected_values[1:-1]))
                # Compare the total column
                self.assertEqual(line_dict_list['horizontal_group_total_data']['no_format'], expected_values[-1])
            else:
                # No total column
                self.assertEqual(len(line_dict_list['columns']), len(expected_values[1:]))
                self.assertEqual(column_values, list(expected_values[1:]))

    def assertHeadersValues(self, headers, expected_headers):
        ''' Helper to compare the headers returned by the _get_table method
        with some expected results.
        An header is a row of columns. Then, headers is a list of list of dictionary.
        :param headers:             The headers to compare.
        :param expected_headers:    The expected headers.
        :return:
        '''
        # Check number of header lines.
        self.assertEqual(len(headers), len(expected_headers))

        for header, expected_header in zip(headers, expected_headers):
            # Check number of columns.
            self.assertEqual(len(header), len(expected_header))

            for i, column in enumerate(header):
                # Check name.
                self.assertEqual(column['name'], self._convert_str_to_date(column['name'], expected_header[i]))

    def assertIdenticalLines(self, reports):
        """Helper to compare report lines with the same `code` across multiple reports.
        The helper checks the lines for similarity on:
        - number of expressions
        - expression label
        - expression engine
        - expression formula
        - expression subformula
        - expression date_scope

        :param reports: (recordset of account.report) The reports to check
        """
        def expression_to_comparable_values(expr):
            return (
                expr.label,
                expr.engine,
                expr.formula,
                expr.subformula,
                expr.date_scope
            )

        if not reports:
            raise UserError('There are no reports to compare.')
        visited_line_codes = set()
        for line in reports.line_ids:
            if not line.code or line.code in visited_line_codes:
                continue
            identical_lines = reports.line_ids.filtered(lambda l: l != line and l.code == line.code)
            if not identical_lines:
                continue
            with self.subTest(line_code=line.code):
                for tested_line in identical_lines:
                    self.assertCountEqual(
                        line.expression_ids.mapped(expression_to_comparable_values),
                        tested_line.expression_ids.mapped(expression_to_comparable_values),
                        (
                            f'The line with code {line.code} from reports "{line.report_id.name}" and '
                            f'"{tested_line.report_id.name}" has different expression values in both reports.'
                        )
                    )
            visited_line_codes.add(line.code)

    def assertLinesValues(self, lines, columns, expected_values, options, currency_map=None, ignore_folded=True):
        ''' Helper to compare the lines returned by the _get_lines method
        with some expected results and ensuring the 'id' key of each line holds a unique value.
        :param lines:               See _get_lines.
        :param columns:             The columns index.
        :param expected_values:     A list of iterables.
        :param options:             The options from the current report.
        :param currency_map:        A map mapping each column_index to some extra options to test the lines:
            - currency:             The currency to be applied on the column.
            - currency_code_index:  The index of the column containing the currency code.
        :param ignore_folded:       Will not filter folded lines when True.
        '''
        if currency_map is None:
            currency_map = {}

        filtered_lines = self._filter_folded_lines(lines) if ignore_folded else lines

        # Compare the table length to see if any line is missing
        self.assertEqual(len(filtered_lines), len(expected_values))

        # Compare cell by cell the current value with the expected one.
        to_compare_list = []
        for i, line in enumerate(filtered_lines):
            compared_values = [[], []]
            for j, index in enumerate(columns):
                if index == 0:
                    current_value = line['name']
                else:
                    # Some lines may not have columns, like title lines. In such case, no values should be provided for these.
                    # Note that the function expect a tuple, so the line still need a comma after the name value.
                    if j > len(expected_values[i]) - 1:
                        break
                    current_value = line['columns'][index-1].get('name', '')
                    current_figure_type = line['columns'][index - 1].get('figure_type', '')

                expected_value = expected_values[i][j]
                currency_data = currency_map.get(index, {})
                used_currency = None
                if 'currency' in currency_data:
                    used_currency = currency_data['currency']
                elif 'currency_code_index' in currency_data:
                    currency_code = line['columns'][currency_data['currency_code_index'] - 1].get('name', '')
                    if currency_code:
                        used_currency = self.env['res.currency'].search([('name', '=', currency_code)], limit=1)
                        assert used_currency, "Currency having name=%s not found." % currency_code
                if not used_currency:
                    used_currency = self.env.company.currency_id

                if type(expected_value) in (int, float) and type(current_value) == str:
                    if current_figure_type and current_figure_type != 'monetary':
                        expected_value = str(expected_value)
                    elif options.get('multi_currency'):
                        expected_value = formatLang(self.env, expected_value, currency_obj=used_currency)
                    else:
                        expected_value = formatLang(self.env, expected_value, digits=used_currency.decimal_places)

                compared_values[0].append(current_value)
                compared_values[1].append(expected_value)

            to_compare_list.append(compared_values)

        errors = []
        for i, to_compare in enumerate(to_compare_list):
            if to_compare[0] != to_compare[1]:
                errors += [
                    "\n==== Differences at index %s ====" % str(i),
                    "Current Values:  %s" % str(to_compare[0]),
                    "Expected Values: %s" % str(to_compare[1]),
                ]

        id_counts = Counter(line['id'] for line in lines)
        duplicate_ids = {k: v for k, v in id_counts.items() if v > 1}
        if duplicate_ids:
            index_to_id = [
                f"index={index:<6} name={line.get('name', 'no line name?!')} \tline_id={line.get('id', 'no line id?!')}"
                for index, line in enumerate(lines)
                if line.get('id', 'no line id?!') in duplicate_ids
            ]
            errors += [
                "\n==== There are lines sharing the same id ====",
                "\n".join(index_to_id)
            ]

        if errors:
            self.fail('\n'.join(errors))

    def _filter_folded_lines(self, lines):
        """ Children lines returned for folded lines (for example, totals below sections) should be ignored when comparing the results
        in assertLinesValues (their parents are folded, so they are not shown anyway). This function returns a filtered version of lines
        list, without the chilren of folded lines.
        """
        filtered_lines = []
        folded_lines = set()
        for line in lines:
            if line.get('parent_id') in folded_lines:
                folded_lines.add(line['id'])
            else:
                if line.get('unfoldable') and not line.get('unfolded'):
                    folded_lines.add(line['id'])
                filtered_lines.append(line)
        return filtered_lines

    def _convert_str_to_date(self, ref, val):
        if isinstance(ref, date) and isinstance(val, str):
            return datetime.strptime(val, '%Y-%m-%d').date()
        return val

    @classmethod
    def _create_tax_report_line(cls, name, report, tag_name=None, parent_line=None, sequence=None, code=None, formula=None):
        """ Creates a tax report line
        """
        create_vals = {
            'name': name,
            'code': code,
            'report_id': report.id,
            'sequence': sequence,
            'expression_ids': [],
        }
        if tag_name and formula:
            raise UserError("Can't use this helper to create a line with both tags and formula")
        if tag_name:
            create_vals['expression_ids'].append(Command.create({
                "label": "balance",
                "engine": "tax_tags",
                "formula": tag_name,
            }))
        if parent_line:
            create_vals['parent_id'] = parent_line.id
        if formula:
            create_vals['expression_ids'].append(Command.create({
                "label": "balance",
                "engine": "aggregation",
                "formula": formula,
            }))

        return cls.env['account.report.line'].create(create_vals)

    @classmethod
    def _get_tag_ids(cls, sign, expressions, company=False):
        """ Helper function to define tag ids for taxes """
        return [(6, 0, cls.env['account.account.tag'].search([
            ('applicability', '=', 'taxes'),
            ('country_id.code', '=', (company or cls.env.company).account_fiscal_country_id.code),
            ('name', 'in', [f"{sign}{f}" for f in expressions.mapped('formula')]),
        ]).ids)]

    @classmethod
    def _get_basic_line_dict_id_from_report_line(cls, report_line):
        """ Computes a full generic id for the provided report line (hence including the one of its parent as prefix), using no markup.
        """
        report = report_line.report_id
        if report_line.parent_id:
            parent_line_id = cls._get_basic_line_dict_id_from_report_line(report_line.parent_id)
            return report._get_generic_line_id(report_line._name, report_line.id, parent_line_id=parent_line_id)

        return report._get_generic_line_id(report_line._name, report_line.id)

    @classmethod
    def _get_basic_line_dict_id_from_report_line_ref(cls, report_line_xmlid):
        """ Same as _get_basic_line_dict_id_from_report_line, but from the line's xmlid, for convenience in the tests.
        """
        return cls._get_basic_line_dict_id_from_report_line(cls.env.ref(report_line_xmlid))

    @classmethod
    def _get_audit_params_from_report_line(cls, options, report_line, report_line_dict, **kwargs):
        return {
            'report_line_id': report_line.id,
            'calling_line_dict_id': report_line_dict['id'],
            'expression_label': 'balance',
            'column_group_key': next(iter(options['column_groups'])),
            **kwargs,
        }

    def _report_compare_with_test_file(self, report, xml_file=None, test_xml=None):
        report_xml = self.get_xml_tree_from_string(report['file_content'])
        if xml_file and not test_xml:
            with file_open(f"{self.test_module}/tests/expected_xmls/{xml_file}", 'rb') as fp:
                test_xml = fp.read()
        test_xml_tree = self.get_xml_tree_from_string(test_xml)
        self.assertXmlTreeEqual(report_xml, test_xml_tree)

    @classmethod
    def _fill_tax_report_line_external_value(cls, target, amount, date):
        cls.env['account.report.external.value'].create({
            'company_id': cls.company_data['company'].id,
            'target_report_expression_id': cls.env.ref(target).id,
            'name': 'Manual value',
            'date': fields.Date.from_string(date),
            'value': amount,
        })

    def _test_xlsx_file(self, file_content, expected_values):
        """ Takes in the binary content of a xlsx file and a dict of expected values.
        It will then parse the file in order to compare the values with the expected ones.
        The expected values dict format is:
        'row_number': ['cell_1_val', 'cell_2_val', ...]

        :param file_content: The binary content of the xlsx file
        :param expected_values: The dict of expected values
        """
        if load_workbook is None:
            raise unittest.SkipTest("openpyxl not available")

        report_file = io.BytesIO(file_content)
        xlsx = load_workbook(filename=report_file, data_only=True)
        sheet = xlsx.worksheets[0]
        sheet_values = list(sheet.values)

        for row, values in expected_values.items():
            row_values = [v if v is not None else '' for v in sheet_values[row]]
            for row_value, expected_value in zip(row_values, values):
                self.assertEqual(row_value, expected_value)
