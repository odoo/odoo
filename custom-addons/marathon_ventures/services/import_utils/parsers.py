# -*- coding: utf-8 -*-
"""Shared CSV/XLS/XLSX readers for tabular imports."""

import csv
import io
import os

import xlrd
from openpyxl import load_workbook
from odoo.exceptions import UserError


def read_tabular_rows(payload, filename, config):
    extension = os.path.splitext(filename or "")[1].lower()
    accepted = config.get("acceptedExtensions", [".csv", ".xls", ".xlsx"])
    if extension not in accepted:
        raise UserError(
            "Unsupported file type. Upload one of: %s" % ", ".join(sorted(accepted))
        )

    header_row = int(config.get("headerRow", 1))
    sheet_index = int(config.get("sheetIndex", 0))

    if extension == ".csv":
        raw = payload.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(raw)))
        return worksheet_rows(rows, header_row=header_row)

    if extension == ".xlsx":
        workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        worksheets = workbook.worksheets
        if sheet_index >= len(worksheets):
            raise UserError("Configured sheetIndex %s is out of range for this workbook." % sheet_index)
        return worksheet_rows(
            worksheets[sheet_index].iter_rows(values_only=True),
            header_row=header_row,
        )

    if extension == ".xls":
        workbook = xlrd.open_workbook(file_contents=payload)
        if sheet_index >= workbook.nsheets:
            raise UserError("Configured sheetIndex %s is out of range for this workbook." % sheet_index)
        sheet = workbook.sheet_by_index(sheet_index)
        rows = [sheet.row_values(idx) for idx in range(sheet.nrows)]
        return worksheet_rows(rows, header_row=header_row)

    raise UserError("Unsupported file type.")


def worksheet_rows(rows, *, header_row=1):
    materialized_rows = [list(row) for row in rows]
    if not materialized_rows:
        return []

    header_index = max(header_row - 1, 0)
    if header_index >= len(materialized_rows):
        raise UserError("Configured headerRow %s is out of range for this file." % header_row)

    header = [
        str(value).strip() if value is not None else ""
        for value in materialized_rows[header_index]
    ]
    normalized_rows = []
    for raw_row in materialized_rows[header_index + 1 :]:
        if not any(value not in (None, "") for value in raw_row):
            continue
        row = {}
        for index, key in enumerate(header):
            row[key] = raw_row[index] if index < len(raw_row) else ""
        normalized_rows.append(row)
    return normalized_rows
