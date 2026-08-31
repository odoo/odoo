# -*- coding: utf-8 -*-
"""Phase 28 - Post Log Import.

Ports the Node.js `handleBundlePostLogData` handler into Odoo. Users
upload a post-log CSV (or XLSX for paid programming) via
Marathon Ventures -> Operations -> Import Post Log; the file is queued
and a background cron matches each row to a mv.schedules record + a
mv.station record, creates the corresponding mv.spot_data rows, and
emails the uploader (plus the fixed program-recipient list) with two
attachments:
  * uploaded-spots.csv - one row per row that landed a Spot Data
  * errors.csv         - one row per row that could not match

Six bundle programs are supported, each with its own column-name
mapping + schedule-matching rules. Common flow lives in
`_process_bundle_common`; per-program specifics (unique matching
paths, discrepancy-schedule cloning for Gray, marathon/paid split
for Univision, hour-split for paid programming) live in the
`_process_<bundle>` methods.
"""
import base64
import csv
import io
import logging
import re
from datetime import date, datetime, timedelta

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


# ============================================================
# Fixed recipient lists (mirroring the Node handler).
# ============================================================
_RECIPIENTS_PAID = ['wendyulloa@mvmediasales.com']
_RECIPIENTS_DEFAULT = [
    'williameccleston@mvmediasales.com',
    'jeffkopczynski@mvmediasales.com',
]


# ============================================================
# Utilities
# ============================================================
def _norm_date(value):
    """Return a YYYY-MM-DD string for whatever date shape came in.
    Accepts str, datetime.date, datetime.datetime. Returns '' if
    it cannot parse."""
    if value is None or value == '':
        return ''
    if isinstance(value, (date, datetime)):
        d = value.date() if isinstance(value, datetime) else value
        return d.strftime('%Y-%m-%d')
    s = str(value).strip()
    if not s:
        return ''
    # Excel serial number sneaking through as string
    if re.fullmatch(r'\d+(\.\d+)?', s):
        try:
            serial = float(s)
            base = date(1899, 12, 30)
            return (base + timedelta(days=serial)).strftime('%Y-%m-%d')
        except (ValueError, OverflowError):
            pass
    for sep in ('-', '/'):
        parts = s.split(sep)
        if len(parts) == 3:
            a, b, c = [p.strip() for p in parts]
            # 2-digit year -> assume 20xx
            if len(c) == 2:
                c = '20' + c
            for fmt in ('%m-%d-%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    normalized = '%s-%s-%s' % (a, b, c) if sep == '-' else '%s/%s/%s' % (a, b, c)
                    return datetime.strptime(normalized, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
    # Last-resort ISO parse
    try:
        return datetime.fromisoformat(s).strftime('%Y-%m-%d')
    except ValueError:
        return ''


def _iso_week_bounds(value):
    """Return (monday, sunday) date strings YYYY-MM-DD for the ISO
    week containing `value`. Empty tuple if the date is unparseable."""
    iso = _norm_date(value)
    if not iso:
        return ('', '')
    d = datetime.strptime(iso, '%Y-%m-%d').date()
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return (monday.strftime('%Y-%m-%d'), sunday.strftime('%Y-%m-%d'))


def _norm_aired_time(value):
    """Return 'H:MM:SS AM/PM' from whatever aired time came in."""
    if value is None or value == '':
        return ''
    if isinstance(value, (datetime,)):
        return value.strftime('%I:%M:%S %p').lstrip('0')
    s = str(value).strip()
    if not s:
        return ''
    # Excel time fraction (0.x day)
    if re.fullmatch(r'0?\.\d+', s):
        try:
            total_sec = round(float(s) * 86400)
            return _seconds_to_ampm(total_sec)
        except ValueError:
            pass
    # 'H:MM' or 'H:MM:SS' with optional AM/PM/XM
    m = re.match(
        r'^\s*(\d{1,2})(?::(\d{1,2}))?(?::(\d{1,2}))?\s*(am|pm|xm)?\s*$',
        s, flags=re.IGNORECASE,
    )
    if not m:
        return s
    hh = int(m.group(1))
    mm = int(m.group(2) or 0)
    ss = int(m.group(3) or 0)
    meridian = (m.group(4) or '').upper()
    if meridian == 'XM':
        meridian = 'AM'
    if meridian:
        h24 = hh % 12
        if meridian == 'PM':
            h24 += 12
    else:
        h24 = hh % 24
    total_sec = h24 * 3600 + mm * 60 + ss
    return _seconds_to_ampm(total_sec)


def _seconds_to_ampm(total_sec):
    total_sec = total_sec % 86400
    h24 = total_sec // 3600
    mm = (total_sec % 3600) // 60
    ss = total_sec % 60
    ap = 'PM' if h24 >= 12 else 'AM'
    h12 = h24 % 12
    if h12 == 0:
        h12 = 12
    return '%d:%02d:%02d %s' % (h12, mm, ss, ap)


def _add_minutes_to_time(aired_time, mins):
    """Shift a 'H:MM:SS AM/PM' string by `mins` minutes."""
    if not aired_time:
        return aired_time
    m = re.match(
        r'^\s*(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?\s*(am|pm|xm)?\s*$',
        aired_time, flags=re.IGNORECASE,
    )
    if not m:
        return aired_time
    hh = int(m.group(1))
    mm = int(m.group(2))
    ss = int(m.group(3) or 0)
    meridian = (m.group(4) or '').upper()
    if meridian == 'XM':
        meridian = 'AM'
    if meridian:
        h24 = hh % 12
        if meridian == 'PM':
            h24 += 12
    else:
        h24 = hh % 24
    total = h24 * 60 + mm + mins
    total = total % 1440
    return _seconds_to_ampm(total * 60 + ss)


def _format_long_form(aired_time_ampm):
    """Round a 'H:MM:SS AM/PM' string to the nearest 30 minutes and
    format as 'A-H:MM' / 'P-H:MM' (matches the Node helper). Used
    by the paid programming composite key."""
    if not aired_time_ampm:
        return ''
    m = re.match(
        r'^(\d{1,2}):(\d{1,2}):(\d{1,2})\s*(AM|PM)$',
        aired_time_ampm.strip(), flags=re.IGNORECASE,
    )
    if not m:
        return ''
    hh = int(m.group(1))
    mm = int(m.group(2))
    ss = int(m.group(3))
    period = m.group(4).upper()
    total_min = hh * 60 + mm + (1 if ss >= 30 else 0)
    total_min = round(total_min / 30) * 30
    new_h = total_min // 60
    new_m = total_min % 60
    if new_h == 0:
        new_h = 12
    if new_h > 12:
        new_h -= 12
    short = 'A' if period == 'AM' else 'P'
    return '%s-%d:%02d' % (short, new_h, new_m)


def _time_to_minutes(time_str):
    if not time_str:
        return 0
    s = str(time_str).strip()
    if ':' in s:
        parts = s.split(':')
        try:
            h = int(parts[0])
            m = int(parts[1])
            return h * 60 + m
        except (ValueError, IndexError):
            return 0
    try:
        return int(s)
    except ValueError:
        return 0


def _length_selection_from_int(n):
    """Turn an integer length (seconds/count) into the mv.spot_data
    Length selection code, e.g. 30 -> 'v_30'. Returns False if not
    a supported value."""
    supported = {5, 10, 15, 30, 45, 60, 75, 90, 120, 180, 300, 3510}
    try:
        n = int(n)
    except (TypeError, ValueError):
        return False
    return 'v_%d' % n if n in supported else False


def _clean_rate(value):
    """Strip currency symbols, return float. '$1,250.00' -> 1250.0."""
    if value is None or value == '':
        return 0.0
    s = str(value)
    s = re.sub(r'[^0-9.]', '', s)
    try:
        return float(s or 0.0)
    except ValueError:
        return 0.0


# ============================================================
# File parsing
# ============================================================
def _parse_upload(filename, raw_bytes):
    """Return a list of dicts (one per row) from a CSV or XLSX blob.
    Delegates on the filename extension."""
    lower = (filename or '').lower()
    if lower.endswith('.xlsx') or lower.endswith('.xls'):
        return _parse_xlsx(raw_bytes)
    return _parse_csv(raw_bytes)


def _parse_csv(raw_bytes):
    text = None
    for enc in ('utf-8-sig', 'utf-8', 'utf-16', 'cp1252', 'latin-1'):
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise UserError(_("Could not decode the uploaded file as text."))
    text = text.replace('\x00', '')
    reader = csv.DictReader(io.StringIO(text, newline=''))
    rows = []
    for r in reader:
        rows.append({(k or '').strip(): (v.strip() if isinstance(v, str) else v)
                     for k, v in r.items()})
    return rows


def _parse_xlsx(raw_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else '' for c in next(it)]
    except StopIteration:
        return []
    rows = []
    for raw in it:
        if all(c is None or c == '' for c in raw):
            continue
        row = {}
        for i, key in enumerate(header):
            if not key:
                continue
            val = raw[i] if i < len(raw) else ''
            if isinstance(val, str):
                val = val.strip()
            row[key] = val
        rows.append(row)
    return rows


# =====================================================================
# Model
# =====================================================================
_PROGRAM_SELECTION = [
    ('tegna',            'Tegna Connect'),
    ('hearst',           'Hearst Unwired / Primary Hearst Connect'),
    ('univision',        'Univision / Unimas Connect'),
    ('american_spirit',  'American Spirit Connect'),
    ('gray',             'Gray (Bounce / Primary / Retro / Telemundo)'),
    ('paid_programming', 'Paid Programming'),
]

_STATE_SELECTION = [
    ('draft',      'Draft'),
    ('queued',     'Queued'),
    ('processing', 'Processing'),
    ('done',       'Done'),
    ('failed',     'Failed'),
]


class MvPostLogImport(models.Model):
    _name = 'mv.post_log_import'
    _description = 'Post Log Import'
    _order = 'create_date desc, id desc'
    _rec_name = 'display_name'

    program = fields.Selection(
        selection=_PROGRAM_SELECTION,
        string='Program',
        required=True,
    )
    upload_file = fields.Binary(
        string='Post Log File',
        required=True,
        attachment=True,
    )
    upload_filename = fields.Char(string='Filename')
    email = fields.Char(
        string='Notification Email',
        required=True,
        default=lambda self: self.env.user.email or '',
        help='The user who receives the results email. Fixed program '
             'recipients (William Eccleston + Jeff Kopczynski for most '
             'programs, Wendy Ulloa for Paid Programming) are copied '
             'automatically.',
    )
    state = fields.Selection(
        selection=_STATE_SELECTION,
        default='draft',
        required=True,
    )
    display_name = fields.Char(compute='_compute_display_name', store=False)
    result_summary = fields.Text(string='Result Summary', readonly=True)
    processed_at = fields.Datetime(string='Processed At', readonly=True)

    @api.depends('program', 'upload_filename', 'create_date')
    def _compute_display_name(self):
        prog_labels = dict(_PROGRAM_SELECTION)
        for rec in self:
            parts = []
            if rec.program:
                parts.append(prog_labels.get(rec.program, rec.program))
            if rec.upload_filename:
                parts.append(rec.upload_filename)
            elif rec.id:
                parts.append('#%s' % rec.id)
            rec.display_name = ' - '.join(parts) if parts else _('New Import')

    # ================================================================
    # Button entry point
    # ================================================================
    def action_queue_import(self):
        """Move the record from draft -> queued and trigger the cron."""
        for rec in self:
            if not rec.upload_file or not rec.program:
                raise UserError(_(
                    "Please pick a Program and upload a file before "
                    "queueing the import."
                ))
            if not rec.email:
                raise UserError(_(
                    "The Notification Email is required."
                ))
            rec.state = 'queued'
        # Trigger the cron so it fires ASAP instead of waiting for its
        # regular tick.
        cron = self.env.ref(
            'marathon_ventures.cron_mv_post_log_import',
            raise_if_not_found=False,
        )
        if cron:
            try:
                cron._trigger()
            except Exception:
                _logger.exception(
                    "post_log_import: cron trigger failed; job stays "
                    "queued and will run on the next tick.",
                )
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Post Log Import'),
                'message': _(
                    "Import queued. You will receive an email at %s "
                    "with the results when processing completes."
                ) % self.email,
                'type': 'success',
                'sticky': False,
            },
        }

    # ================================================================
    # Cron-driven processing
    # ================================================================
    @api.model
    def _cron_process_queued(self):
        """Called by ir.cron. Picks up queued jobs and runs them one
        at a time. Each job commits its own transaction so a failure
        on one doesn't block the next."""
        jobs = self.search([('state', '=', 'queued')], order='create_date asc')
        for job in jobs:
            try:
                job._process_one()
                self.env.cr.commit()
            except Exception:
                _logger.exception(
                    "post_log_import: job id=%s failed", job.id,
                )
                self.env.cr.rollback()
                job.write({
                    'state': 'failed',
                    'result_summary': 'Unexpected error - see server log.',
                    'processed_at': fields.Datetime.now(),
                })
                self.env.cr.commit()

    def _process_one(self):
        self.ensure_one()
        self.write({'state': 'processing'})
        # Materialize the upload from the attachment column.
        raw = base64.b64decode(self.upload_file) if self.upload_file else b''
        if not raw:
            raise UserError(_("The uploaded file is empty."))
        rows = _parse_upload(self.upload_filename, raw)
        if not rows:
            raise UserError(_("The uploaded file has no data rows."))

        dispatch = {
            'tegna':            self._process_tegna,
            'hearst':           self._process_hearst,
            'univision':        self._process_univision,
            'american_spirit':  self._process_american_spirit,
            'gray':             self._process_gray,
            'paid_programming': self._process_paid_programming,
        }
        handler = dispatch.get(self.program)
        if not handler:
            raise UserError(_(
                "No handler is wired for program '%s'."
            ) % self.program)
        success, errors = handler(rows)

        self.write({
            'state': 'done',
            'processed_at': fields.Datetime.now(),
            'result_summary': _(
                "Processed %(total)d row(s): %(ok)d created, "
                "%(err)d error(s)."
            ) % {
                'total': len(success) + len(errors),
                'ok': len(success),
                'err': len(errors),
            },
        })
        self._send_result_email(success, errors)

    # ================================================================
    # Result email
    # ================================================================
    def _send_result_email(self, success_rows, error_rows):
        self.ensure_one()
        prog_labels = dict(_PROGRAM_SELECTION)
        prog_label = prog_labels.get(self.program, self.program)

        recipients = {self.email}
        if self.program == 'paid_programming':
            recipients.update(_RECIPIENTS_PAID)
        else:
            recipients.update(_RECIPIENTS_DEFAULT)
        recipients.discard('')

        # Build CSVs. We ALWAYS emit both files - even when a list is
        # empty we ship a header-only file so the user has a receipt of
        # what happened + a clear "0 rows" signal, and so the email
        # always carries visible attachments.
        def _to_csv(rows, fallback_header):
            fieldnames = list(rows[0].keys()) if rows else fallback_header
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
            return buf.getvalue()

        _SUCCESS_FALLBACK = [
            'Schedule__c', 'Raycom_Order_Number__c', 'Air_Date__c',
            'Air_Time__c', 'Length__c', 'Spot_Rate__c', 'ISCI__c',
            'Station__c', 'Station Name', 'Result', 'Result Details',
            'Spot Data ID',
        ]
        _ERROR_FALLBACK = _SUCCESS_FALLBACK
        success_csv = _to_csv(success_rows, _SUCCESS_FALLBACK)
        errors_csv  = _to_csv(error_rows,   _ERROR_FALLBACK)

        subject = (
            "Post Log Upload Complete: %d error(s), %d spot(s) created "
            "for %s (%s)"
        ) % (
            len(error_rows), len(success_rows),
            self.upload_filename or '', prog_label,
        )
        body = _(
            "Your post log upload of file %(fn)s completed.\n"
            "%(ok)d rows were successfully uploaded as spot records "
            "and matched to their schedules (see attached "
            "uploaded-spots.csv).\n"
            "%(err)d rows were not processed due to errors (see "
            "attached errors.csv)."
        ) % {
            'fn': self.upload_filename or '',
            'ok': len(success_rows),
            'err': len(error_rows),
        }

        Mail = self.env['mail.mail'].sudo()
        Attachment = self.env['ir.attachment'].sudo()

        # Anchor the attachments on this job record so they don't get
        # garbage-collected before the outgoing mail worker picks them
        # up. auto_delete on the mail record won't remove them because
        # they're now owned by mv.post_log_import.
        att_success = Attachment.create({
            'name': 'uploaded-spots.csv',
            'datas': base64.b64encode(success_csv.encode('utf-8')),
            'mimetype': 'text/csv',
            'type': 'binary',
            'res_model': self._name,
            'res_id': self.id,
        })
        att_errors = Attachment.create({
            'name': 'errors.csv',
            'datas': base64.b64encode(errors_csv.encode('utf-8')),
            'mimetype': 'text/csv',
            'type': 'binary',
            'res_model': self._name,
            'res_id': self.id,
        })
        att_ids = [att_success.id, att_errors.id]
        _logger.info(
            "post_log_import: job id=%s emailing to %s with attachment "
            "ids=%s (uploaded=%d bytes, errors=%d bytes)",
            self.id, sorted(recipients), att_ids,
            len(success_csv or ''), len(errors_csv or ''),
        )

        mail = Mail.create({
            'subject': subject,
            'body_html': '<pre>%s</pre>' % body,
            'email_to': ','.join(sorted(recipients)),
            'attachment_ids': [(6, 0, att_ids)],
            # We keep auto_delete = False so if the SMTP server is slow
            # or the message ends up in the outgoing queue, the mail
            # record + its attachment linkage stay intact until the
            # worker picks it up.
            'auto_delete': False,
        })
        try:
            mail.send(raise_exception=False)
        except Exception:
            _logger.exception(
                "post_log_import: send failed for job id=%s",
                self.id,
            )

    # ================================================================
    # Common helpers used by all handlers
    # ================================================================
    def _station_map(self, active_only=False):
        """Return {call_letters: station_id}. Owner_Call_Letters also
        maps to the same station id."""
        Station = self.env['mv.station'].sudo()
        domain = []
        if active_only and 'active_station' in Station._fields:
            domain.append(('active_station', '=', True))
        m = {}
        for st in Station.search(domain):
            call = st.call_letters or ''
            if call:
                # SF stored slash-separated aliases in the same field.
                for token in [t.strip() for t in call.split('/') if t.strip()]:
                    m[token] = st.id
            owner = st.owner_call_letters or ''
            if owner:
                m[owner] = st.id
        return m

    def _schedule_maps(self, domain):
        """Return two dicts covering the schedule set matched by
        `domain`:
          deal_name -> schedule_id
          isci      -> schedule_id
        `isci_code` may hold multiple slash-separated codes; every
        code becomes its own key."""
        Schedule = self.env['mv.schedules'].sudo()
        deal_map = {}
        isci_map = {}
        for sc in Schedule.search(domain):
            sid = sc.id
            deal_name = sc.deal_parent.name if sc.deal_parent else ''
            if deal_name:
                deal_map[deal_name] = sid
            isci = sc.isci_code or ''
            if isci:
                for token in [t.strip() for t in isci.split('/') if t.strip()]:
                    isci_map[token] = sid
        return deal_map, isci_map

    def _insert_spot_data_batch(self, records):
        """Batch-create mv.spot_data rows. Returns two aligned lists:
          created_ids : [record_id or False for each input]
          errors      : ['' or reason string for each input]
        """
        SpotData = self.env['mv.spot_data'].sudo()
        created_ids = []
        errors = []
        for vals in records:
            try:
                rec = SpotData.create(vals)
                created_ids.append(rec.id)
                errors.append('')
            except Exception as e:
                created_ids.append(False)
                errors.append(str(e))
        return created_ids, errors

    # ================================================================
    # Per-program processors
    # ------------------------------------------------------------
    # Each processor returns (success_rows, error_rows) - lists of
    # dicts written to the emailed CSVs.
    # ================================================================
    # ---- Tegna ---------------------------------------------------
    def _process_tegna(self, rows):
        first_date = rows[0].get('Broadcast Date', '')
        wk_start, wk_end = _iso_week_bounds(first_date)
        if not wk_start:
            raise UserError(_("Broadcast Date missing in the first row."))

        domain = [
            ('deal_parent.program.name', 'ilike', 'tegna'),
            ('status', '=', 'sold'),
            ('deal_parent.pi', '=', False),
            ('deal_parent.status', '!=', 'canceled'),
            ('week', '>=', wk_start),
            ('week', '<=', wk_end),
        ]
        deal_map, isci_map = self._schedule_maps(domain)
        station_map = self._station_map()

        return self._process_bundle_common(
            rows,
            date_col='Broadcast Date',
            order_col='Order Number',
            time_col='Airing Time',
            length_col='Length',
            rate_col='Rate',
            isci_col='ISCI',
            station_col='Station',
            deal_lookup=lambda row: self._extract_tegna_deal_id(row),
            deal_map=deal_map,
            isci_map=isci_map,
            station_map=station_map,
            station_label_col='Station',
            extra_result_cols={'Estimate': lambda row: row.get('Estimate', '')},
        )

    def _extract_tegna_deal_id(self, row):
        est = (row.get('Estimate') or '').strip()
        if ':' in est:
            return est.split(':', 1)[1].strip()
        return ''

    # ---- Hearst --------------------------------------------------
    def _process_hearst(self, rows):
        first_date = rows[0].get('Broadcast Date', '')
        wk_start, wk_end = _iso_week_bounds(first_date)
        if not wk_start:
            raise UserError(_("Broadcast Date missing in the first row."))

        domain = [
            '|',
            ('deal_parent.program.name', '=', 'Hearst Unwired'),
            ('deal_parent.program.name', '=', 'Primary Hearst Connect'),
            ('status', '=', 'sold'),
            ('deal_parent.pi', '=', False),
            ('deal_parent.status', '!=', 'canceled'),
            ('week', '>=', wk_start),
            ('week', '<=', wk_end),
        ]
        deal_map, isci_map = self._schedule_maps(domain)
        station_map = self._station_map()

        return self._process_bundle_common(
            rows,
            date_col='Broadcast Date',
            order_col='Order Id',
            time_col='Hit Time',
            length_col='Length',
            rate_col='Rate',
            isci_col='ISCI',
            station_col='Media Outlet Name',
            deal_lookup=lambda row: (row.get('Alt Order Id') or '').strip(),
            deal_map=deal_map,
            isci_map=isci_map,
            station_map=station_map,
            station_label_col='Media Outlet Name',
            length_transform=_time_to_minutes,
        )

    # ---- American Spirit ----------------------------------------
    def _process_american_spirit(self, rows):
        first_date = rows[0].get('Date', '')
        wk_start, wk_end = _iso_week_bounds(first_date)
        if not wk_start:
            raise UserError(_("Air Date missing in the first row."))

        domain = [
            ('deal_parent.program.name', '=', 'American Spirit Connect'),
            ('status', '=', 'sold'),
            ('deal_parent.pi', '=', False),
            ('deal_parent.status', '!=', 'canceled'),
            ('week', '>=', wk_start),
            ('week', '<=', wk_end),
        ]
        deal_map, isci_map = self._schedule_maps(domain)
        station_map = self._station_map(active_only=True)

        return self._process_bundle_common(
            rows,
            date_col='Date',
            order_col='Order',
            time_col='Air Time',
            length_col='Length',
            rate_col='Rate (Ext)',
            isci_col='Ad-ID',
            station_col='Property',
            deal_lookup=lambda row: (row.get('Estimate #') or '').strip(),
            deal_map=deal_map,
            isci_map=isci_map,
            station_map=station_map,
            station_label_col='Property',
            length_transform=_time_to_minutes,
        )

    # ---- Univision (marathon + paid split) ----------------------
    def _process_univision(self, rows):
        first_date = rows[0].get('Air Date', '')
        wk_start, wk_end = _iso_week_bounds(first_date)
        if not wk_start:
            raise UserError(_("Air Date missing in the first row."))

        marathon_domain = [
            '|',
            ('deal_parent.program.name', '=', 'Univision Connect'),
            ('deal_parent.program.name', '=', 'Unimas Connect'),
            ('deal_parent.dealaccount', '=', 'Marathon Ventures'),
            ('deal_parent.brands.name', '!=', 'availability'),
            ('status', '=', 'sold'),
            ('deal_parent.pi', '=', False),
            ('week', '>=', wk_start),
            ('week', '<=', wk_end),
        ]
        paid_domain = [
            '|',
            ('deal_parent.program.name', '=', 'Univision Connect'),
            ('deal_parent.program.name', '=', 'Unimas Connect'),
            ('deal_parent.dealaccount', '!=', 'Marathon Ventures'),
            ('deal_parent.brands.name', '!=', 'availability'),
            ('status', '=', 'sold'),
            ('deal_parent.pi', '=', False),
            ('week', '>=', wk_start),
            ('week', '<=', wk_end),
        ]
        marathon_deal_map, marathon_isci_map = self._schedule_maps(marathon_domain)
        paid_deal_map, paid_isci_map = self._schedule_maps(paid_domain)
        station_map = self._station_map()

        # Two passes: marathon set, then paid set. Concatenate results.
        marathon_ok, marathon_err = self._process_bundle_common(
            rows,
            date_col='Air Date',
            order_col='Order #',
            time_col='Air Time',
            length_col='Length',
            rate_col='Rate (Ext)',
            isci_col='Ad-ID',
            station_col='Property',
            deal_lookup=lambda row: (row.get('Alt Order #') or '').strip(),
            deal_map=marathon_deal_map,
            isci_map=marathon_isci_map,
            station_map=station_map,
            station_label_col='Property',
            length_transform=_time_to_minutes,
        )
        paid_ok, paid_err = self._process_bundle_common(
            rows,
            date_col='Air Date',
            order_col='Order #',
            time_col='Air Time',
            length_col='Length',
            rate_col='Rate (Ext)',
            isci_col='Ad-ID',
            station_col='Property',
            deal_lookup=lambda row: (row.get('Alt Order #') or '').strip(),
            deal_map=paid_deal_map,
            isci_map=paid_isci_map,
            station_map=station_map,
            station_label_col='Property',
            length_transform=_time_to_minutes,
        )
        # Tag each row with the branch so the CSV shows which side ran.
        for r in marathon_ok:
            r['Branch'] = 'Marathon'
        for r in marathon_err:
            r['Branch'] = 'Marathon'
        for r in paid_ok:
            r['Branch'] = 'Paid'
        for r in paid_err:
            r['Branch'] = 'Paid'
        return marathon_ok + paid_ok, marathon_err + paid_err

    # ---- Gray (with discrepancy schedule cloning) ---------------
    _GRAY_DELETE_CALL_LETTERS = {
        'EDBD', 'HXTX', 'KAUZ', 'KVHP', 'MVHP',
        'WDBD', 'WLOO', 'WSFX', 'WXTX',
    }

    def _process_gray(self, rows):
        rows = [r for r in rows
                if (r.get('Media Outlet Name') or '') not in self._GRAY_DELETE_CALL_LETTERS]
        if not rows:
            return [], []
        first_date = rows[0].get('Broadcast Date', '')
        wk_start, wk_end = _iso_week_bounds(first_date)
        if not wk_start:
            raise UserError(_("Broadcast Date missing in the first row."))

        sold_domain = [
            ('deal_parent.program.name', 'ilike', 'gray'),
            ('status', '=', 'sold'),
            ('deal_parent.pi', '=', False),
            ('deal_parent.status', '!=', 'canceled'),
            ('deal_parent.brands.name', '!=', 'availability'),
            ('week', '>=', wk_start),
            ('week', '<=', wk_end),
        ]
        cancelled_domain = [
            ('deal_parent.program.name', 'ilike', 'gray'),
            ('status', '=', 'canceled'),
            ('deal_parent.pi', '=', False),
            ('deal_parent.brands.name', '!=', 'availability'),
            ('week', '>=', wk_start),
            ('week', '<=', wk_end),
        ]
        deal_map, isci_map = self._schedule_maps(sold_domain)
        cancel_deal_map, _cancel_isci = self._schedule_maps(cancelled_domain)
        station_map = self._station_map()

        # First pass: clone any cancelled schedule whose deal appears
        # in the file (Gray discrepancy-schedule behaviour).
        to_clone = {}  # cancelled_sched_id -> deal_id_from_file
        for row in rows:
            deal_id = (row.get('Alt Order Id') or '').strip()
            if deal_id in deal_map:
                continue
            if deal_id in cancel_deal_map:
                to_clone[cancel_deal_map[deal_id]] = deal_id

        Schedule = self.env['mv.schedules'].sudo()
        if to_clone:
            for sched_id, deal_id in to_clone.items():
                orig = Schedule.browse(sched_id)
                if not orig.exists():
                    continue
                clone_vals = self._gray_clone_vals(orig)
                new_sc = Schedule.create(clone_vals)
                deal_map[deal_id] = new_sc.id
                if new_sc.isci_code:
                    for token in [t.strip() for t in new_sc.isci_code.split('/') if t.strip()]:
                        isci_map[token] = new_sc.id

        return self._process_bundle_common(
            rows,
            date_col='Broadcast Date',
            order_col='Order Id',
            time_col='Hit Time',
            length_col='Length',
            rate_col='Rate',
            isci_col='ISCI',
            station_col='Media Outlet Name',
            deal_lookup=lambda row: (row.get('Alt Order Id') or '').strip(),
            deal_map=deal_map,
            isci_map=isci_map,
            station_map=station_map,
            station_label_col='Media Outlet Name',
            length_transform=_time_to_minutes,
            time_transform=lambda t: (t or '').replace('xm', 'AM').replace('XM', 'AM')
                                              .replace('Xm', 'AM').replace('xM', 'AM'),
        )

    def _gray_clone_vals(self, orig):
        """Build a create-vals dict from an existing schedule, dropping
        the fields that should not be cloned + setting the discrepancy
        overrides."""
        blacklist = {
            'id', 'display_name', 'create_uid', 'create_date',
            'write_uid', 'write_date', '__last_update', 'name',
            'sf_external_id',
        }
        vals = {}
        for fname, field in orig._fields.items():
            if fname in blacklist:
                continue
            if field.compute and not field.store:
                continue
            if field.related:
                continue
            if field.type in ('one2many',):
                continue
            v = orig[fname]
            if field.type == 'many2one':
                vals[fname] = v.id if v else False
            elif field.type == 'many2many':
                vals[fname] = [(6, 0, v.ids)]
            else:
                vals[fname] = v
        # Discrepancy overrides
        if 'discrepancy_bundle' in orig._fields:
            vals['discrepancy_bundle'] = True
        if 'error_reason' in orig._fields:
            vals['error_reason'] = 'Network - Cancelled Spots Aired'
        vals['rate'] = 0.0
        vals['status'] = 'sold'
        return vals

    # ---- Paid Programming (xlsx path + composite key) -----------
    def _process_paid_programming(self, rows):
        # Row normalization (Property fallback + CHANNEL_MAP + hour split)
        rows = self._paid_normalize_rows(rows)
        rows = self._paid_split_hour_schedules(rows)
        if not rows:
            raise UserError(_("The upload had no usable rows after normalisation."))
        first_date = rows[0].get('Air Date', '')
        wk_start, wk_end = _iso_week_bounds(first_date)
        if not wk_start:
            raise UserError(_("Air Date missing in the first row."))

        domain = [
            ('deal_parent.program.name', 'in', [
                'Ion Television PP', 'Grit PP', 'Laff PP',
                'Bounce PP', 'Outlaw PP',
            ]),
            ('lf_week', '>=', wk_start),
            ('lf_week', '<=', wk_end),
            ('status', '=', 'sold'),
        ]
        Schedule = self.env['mv.schedules'].sudo()
        scheds = Schedule.search(domain)

        # Composite key: program|account|long_form|week|rate  (lower-case)
        key_to_id = {}
        agencies = set()
        for sc in scheds:
            deal_account = (sc.deal_parent.dealaccount or '').strip()
            final_acct = 'Cannella Response Television' \
                if deal_account.lower() == 'cannella media' else deal_account
            key = '|'.join([
                (sc.deal_parent.program.name if sc.deal_parent.program else '').strip().lower(),
                final_acct.strip().lower(),
                (sc.long_form or '').strip(),
                (sc.week.isoformat() if sc.week else '').strip(),
                str(sc.rate or '').strip(),
            ])
            key_to_id[key] = sc.id
            if final_acct:
                agencies.add(final_acct)

        success_rows, error_rows = [], []
        to_insert = []
        for row in rows:
            row_agency = row.get('Agency') or ''
            agency = next(
                (a for a in agencies
                 if a and (a in row_agency or row_agency in a)),
                row_agency,
            )
            key = '|'.join([
                (row.get('Property') or '').strip().lower(),
                (agency or '').strip().lower(),
                (row.get('Long Form') or '').strip(),
                (row.get('Week') or '').strip(),
                str(row.get('Rate') or '').strip(),
            ])
            sched_id = key_to_id.get(key)
            aired_length = row.get('Aired Length', '')
            total_seconds = 0
            if aired_length and ':' in str(aired_length):
                try:
                    parts = str(aired_length).split(':')
                    total_seconds = int(parts[0]) * 60 + int(parts[1])
                except ValueError:
                    total_seconds = 0
            length_sel = _length_selection_from_int(total_seconds)
            vals = {
                'schedule': sched_id or False,
                'raycom_order_number': row.get('Deal/Order #', ''),
                'air_date': _norm_date(row.get('Air Date', '')) or False,
                'air_time': row.get('Aired Time', ''),
                'length': length_sel,
                'spot_rate': _clean_rate(row.get('Rate', 0)),
                'isci': row.get('Aired Ad-ID', ''),
                'x800': row.get('Field 1', ''),
                'commercial_title': row.get('Order Product Description', ''),
            }
            if not sched_id:
                error_rows.append({
                    **row,
                    'Status': 'Failed',
                    'Reason': 'No schedule matched',
                    'Spot Data ID': '',
                })
                continue
            to_insert.append((vals, row))

        created_ids, errors = self._insert_spot_data_batch(
            [v for (v, _r) in to_insert],
        )
        for i, (vals, row) in enumerate(to_insert):
            if created_ids[i]:
                success_rows.append({
                    **row,
                    'Status': 'Uploaded',
                    'Reason': '',
                    'Spot Data ID': created_ids[i],
                })
            else:
                error_rows.append({
                    **row,
                    'Status': 'Error',
                    'Reason': errors[i] or 'Insert failed',
                    'Spot Data ID': '',
                })
        return success_rows, error_rows

    _PAID_CHANNEL_MAP = {
        'LAF': 'Laff PP',
        'GRT': 'GRIT PP',
        'ION': 'ION Television PP',
        'BTV': 'Bounce PP',
    }
    _PAID_HOUR_SPLIT_LABEL = 'hour split schedule'

    def _paid_normalize_rows(self, rows):
        out = []
        for r in rows:
            row = {(k.strip() if isinstance(k, str) else k): v for k, v in r.items()}
            row['Property'] = row.get('Property') or row.get('Channels (Placed)') or ''
            if not (row.get('Property') or '').strip():
                continue
            row['Property'] = self._PAID_CHANNEL_MAP.get(row['Property'], row['Property'])
            row['Air Date']  = _norm_date(row.get('Air Date'))
            row['Aired Time'] = _norm_aired_time(row.get('Aired Time'))
            row['Week']       = row['Air Date']  # Air Date already ISO
            row['Long Form']  = _format_long_form(row['Aired Time'])
            out.append(row)
        return out

    def _paid_split_hour_schedules(self, rows):
        result = []
        for row in rows:
            agency = (row.get('Agency') or '').strip()
            desc = str(row.get('Order Product Description') or '')
            aired_length = str(row.get('Aired Length') or '').strip()
            if (
                agency == 'Apex Media - Arizona'
                and re.search(r'campmeeting|uncommon blessing', desc, re.IGNORECASE)
                and aired_length.startswith('58:30')
            ):
                numeric_rate = _clean_rate(row.get('Rate', ''))
                half_rate = numeric_rate / 2
                first = {**row, 'Rate': half_rate, 'Program': self._PAID_HOUR_SPLIT_LABEL}
                second = {
                    **row, 'Rate': half_rate,
                    'Program': self._PAID_HOUR_SPLIT_LABEL,
                    'Aired Time': _add_minutes_to_time(row.get('Aired Time'), 30),
                }
                result.append(first)
                result.append(second)
            else:
                result.append(row)
        return result

    # ================================================================
    # Common processor used by tegna / hearst / american_spirit /
    # gray / univision. Paid programming has its own path above.
    # ================================================================
    def _process_bundle_common(self, rows,
                               date_col, order_col, time_col, length_col,
                               rate_col, isci_col, station_col,
                               deal_lookup,
                               deal_map, isci_map, station_map,
                               station_label_col,
                               length_transform=None,
                               time_transform=None,
                               extra_result_cols=None):
        success_rows, error_rows = [], []
        to_insert = []
        for row in rows:
            deal_id = deal_lookup(row) or ''
            row_isci = (row.get(isci_col) or '').strip()
            sched_id = deal_map.get(deal_id) or isci_map.get(row_isci)
            station_id = station_map.get((row.get(station_col) or '').strip())

            raw_length = row.get(length_col, '')
            length_value = length_transform(raw_length) if length_transform else raw_length
            length_sel = _length_selection_from_int(length_value)

            air_time = row.get(time_col, '')
            if time_transform:
                air_time = time_transform(air_time)

            base_result = {
                'Schedule__c': sched_id or '',
                'Raycom_Order_Number__c': row.get(order_col, ''),
                'Air_Date__c': _norm_date(row.get(date_col, '')),
                'Air_Time__c': air_time,
                'Length__c': length_value,
                'Spot_Rate__c': _clean_rate(row.get(rate_col, 0)),
                'ISCI__c': row_isci,
                'Station__c': station_id or '',
                'Station Name': row.get(station_label_col, '') or '',
                'Spot Data ID': '',
            }
            for k, fn in (extra_result_cols or {}).items():
                base_result[k] = fn(row)

            if not sched_id:
                error_rows.append({
                    **base_result,
                    'Result': 'Error',
                    'Result Details': 'No schedule match',
                })
                continue
            if not station_id:
                error_rows.append({
                    **base_result,
                    'Result': 'Error',
                    'Result Details': 'Station Not Found',
                })
                continue
            vals = {
                'schedule': sched_id,
                'raycom_order_number': row.get(order_col, ''),
                'air_date': _norm_date(row.get(date_col, '')) or False,
                'air_time': air_time,
                'length': length_sel,
                'spot_rate': _clean_rate(row.get(rate_col, 0)),
                'isci': row_isci,
                'station': station_id,
            }
            to_insert.append((vals, base_result, row))

        created_ids, errors = self._insert_spot_data_batch(
            [v for (v, _b, _r) in to_insert],
        )
        for i, (vals, base_result, row) in enumerate(to_insert):
            if created_ids[i]:
                success_rows.append({
                    **base_result,
                    'Result': 'Uploaded',
                    'Result Details': '',
                    'Spot Data ID': created_ids[i],
                })
            else:
                error_rows.append({
                    **base_result,
                    'Result': 'Error',
                    'Result Details': errors[i] or 'Insert failed',
                })
        return success_rows, error_rows
